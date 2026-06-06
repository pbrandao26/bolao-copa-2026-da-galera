from pathlib import Path
import random
from openpyxl import load_workbook

template_path = Path(
    r"C:\Users\pedro\Projetos\bolao-copa-2026-da-galera\apostas\Bolao_Copa2026_Pedro Brandão.xlsx"
)

output_dir = Path(
    r"C:\Users\pedro\Projetos\bolao-copa-2026-da-galera\apostas"
)

nomes = [
    "João Silva",
    "Maria Souza",
    "Carlos Lima",
    "Ana Costa",
    "Bruno Rocha",
    "Lucas Pereira",
    "Fernanda Alves",
    "Rafael Gomes",
    "Juliana Martins",
    "Pedro Santos",
    "Mariana Ferreira",
    "Gustavo Ribeiro",
    "Camila Barros",
    "Thiago Almeida",
    "Beatriz Nunes",
    "Rodrigo Teixeira",
    "Larissa Carvalho",
    "Felipe Mendes",
    "Amanda Castro",
    "Daniel Moreira",
    "Patrícia Lopes",
    "Eduardo Araújo",
    "Renata Cardoso",
    "Marcelo Duarte",
    "Isabela Freitas",
    "Vinicius Oliveira",
    "Carolina Batista",
    "André Farias",
    "Letícia Ramos",
    "Henrique Moura",
]

artilheiros = [
    "Kylian Mbappé",
    "Erling Haaland",
    "Vinícius Júnior",
    "Harry Kane",
    "Lionel Messi",
    "Lautaro Martínez",
    "Julián Álvarez",
    "Cristiano Ronaldo",
    "Viktor Gyökeres",
    "Mohamed Salah",
]

for nome in nomes:

    wb = load_workbook(template_path)

    ws_grupos = wb["Apostas - Grupos"]
    ws_bonus = wb["Apostas - Bonus"]

    # Nome participante
    ws_grupos["A5"] = nome

    # 72 jogos = 144 células de placar
    for col in range(2, 146):
        ws_grupos.cell(
            row=5,
            column=col
        ).value = random.randint(0, 5)

    # Artilheiro aleatório
    ws_bonus["B5"] = random.choice(artilheiros)

    output_name = f"Bolao_Copa2026_{nome}.xlsx"

    output_path = output_dir / output_name

    wb.save(output_path)
    wb.close()

print("30 planilhas criadas com sucesso.")