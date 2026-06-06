#!/usr/bin/env python3
"""
Bolão Copa do Mundo 2026 — Versão Fase de Grupos
Dashboard simplificado · só fase de grupos + bônus artilheiro (20 pts)
"""

import streamlit as st
import plotly.graph_objects as go
from openpyxl import load_workbook
import os, re, base64, hashlib, pickle
from concurrent.futures import ThreadPoolExecutor
from collections import OrderedDict
from datetime import date
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

try:
    _PWD = st.secrets["APP_PASSWORD"]
except Exception:
    _PWD = os.getenv("APP_PASSWORD")

st.set_page_config(
    page_title="Bolão Copa 2026 · Fase de Grupos",
    page_icon="⚽",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:ital,wght@0,300;0,400;0,600;0,700;0,800;0,900;1,400&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', 'Effra', 'Gotham Book', sans-serif !important;
}

.block-container { padding-top: 1.2rem !important; max-width: 1400px; }

.hero {
    background: linear-gradient(135deg, #0D2B40 0%, #123A56 50%, #0D8587 100%);
    border-radius: 16px; padding: 28px 40px; margin-bottom: 20px;
    position: relative; overflow: hidden;
    box-shadow: 0 6px 28px rgba(13,43,64,.35);
}
.hero::after {
    content: ""; position: absolute; right: -20px; top: -20px;
    width: 200px; height: 200px; border-radius: 50%;
    background: rgba(214,184,100,.12);
}
.hero-title {
    font-size: 1.85rem; font-weight: 900; color: #FFFFFF !important;
    margin: 0; letter-spacing: -.5px;
}
.hero-sub { font-size: .88rem; color: rgba(255,255,255,.72) !important; margin-top: 5px; }
.hero-badge {
    display: inline-block; background: rgba(214,184,100,.25);
    border: 1px solid rgba(214,184,100,.5); border-radius: 20px;
    padding: 3px 12px; font-size: .75rem; color: #D6B864 !important; margin-top: 10px;
}

.mc {
    border: 1px solid rgba(128,128,128,.15);
    border-radius: 12px; padding: 14px 16px; text-align: center;
}
.mc-v { font-size: 1.7rem; font-weight: 900; color: #D6B864 !important; line-height: 1; }
.mc-l { font-size: .63rem; color: inherit; opacity: .6; text-transform: uppercase;
         letter-spacing: 1px; margin-top: 4px; }

.rc {
    border-radius: 11px; padding: 12px 16px; margin-bottom: 6px;
    display: flex; align-items: center; gap: 12px;
    border: 1px solid rgba(128,128,128,.15);
}
.rc1 { border-left: 4px solid #D6B864; }
.rc2 { border-left: 4px solid #7F7F7F; }
.rc3 { border-left: 4px solid #DC884A; }
.rcN { border-left: 4px solid #123A56; }
.rc-name { font-size: .94rem; font-weight: 700; flex: 1; }
.rc-sub  { font-size: .67rem; opacity: .6; margin-top: 2px; }
.rc-pts  { font-size: 1.55rem; font-weight: 900; color: #D6B864 !important; }
.rc-pl   { font-size: .62rem; color: #5C5F62; text-align: right; }
.bar-bg  { height: 3px; border-radius: 3px; background: rgba(18,58,86,.12); margin-top: 5px; }
.bar-fg  { height: 3px; border-radius: 3px; background: linear-gradient(90deg, #123A56, #D6B864); }

.gb {
    border: 1px solid rgba(128,128,128,.12); border-radius: 11px;
    padding: 11px 12px; margin-bottom: 9px;
}
.gb-hdr { font-weight: 800; font-size: .88rem; margin-bottom: 6px; }
.gt { width: 100%; border-collapse: collapse; font-size: .76rem; }
.gt th { font-weight: 700; text-align: center; padding: 4px 4px;
          background: #0D2B40; color: #FFFFFF !important;
          border-bottom: 2px solid rgba(214,184,100,.3); }
.gt td { padding: 4px 4px; text-align: center; color: inherit; }
.gt td.nm { text-align: left; font-weight: 600; white-space: nowrap; }
.dot { display: inline-block; width: 7px; height: 7px; border-radius: 50%;
       margin-right: 4px; vertical-align: middle; }

.b5 { background: #14532D; color: #86EFAC; border-radius: 4px; padding: 1px 7px; font-size: .7rem; font-weight: 700; }
.b3 { background: #0D4040; color: #5EEAD4; border-radius: 4px; padding: 1px 7px; font-size: .7rem; font-weight: 700; }
.b2 { background: #7C2D12; color: #FED7AA; border-radius: 4px; padding: 1px 7px; font-size: .7rem; font-weight: 700; }
.b0 { background: #7F1D1D; color: #FECACA; border-radius: 4px; padding: 1px 7px; font-size: .7rem; font-weight: 700; }
.bN { border-radius: 4px; padding: 1px 7px; font-size: .7rem; opacity: .5; }

.mr {
    border: 1px solid rgba(128,128,128,.10); border-radius: 8px;
    padding: 6px 10px; margin-bottom: 3px;
    display: flex; align-items: center; gap: 8px; font-size: .78rem;
}
.mr-t { flex: 1; font-weight: 600; }
.mr-s { font-size: .73rem; color: #5C5F62; white-space: nowrap; }

.mm-tbl { border-collapse: collapse; font-size: .78rem; width: 100%; min-width: 500px; }
.mm-tbl th {
    background: #0D2B40; color: #FFFFFF !important;
    padding: 7px 9px; text-align: center; border: 1px solid rgba(255,255,255,.1);
    white-space: nowrap; font-weight: 700;
}
.mm-tbl td { padding: 6px 9px; text-align: center; white-space: nowrap;
             color: inherit; border: 1px solid rgba(128,128,128,.10); }
.mm-tbl tr:nth-child(even) td { background: rgba(18,58,86,.04); }

.bc {
    border: 1px solid rgba(128,128,128,.12); border-radius: 11px;
    padding: 16px; text-align: center;
}
.bc-lbl { font-size: .63rem; color: #5C5F62 !important; text-transform: uppercase;
           letter-spacing: 1.1px; margin-bottom: 5px; }
.bc-ico  { font-size: 1.5rem; }
.bc-bet  { font-size: .88rem; font-weight: 600; margin: 4px 0 2px; }
.bc-real { font-size: .76rem; opacity: .6; }
.bc-pts  { font-size: 1.35rem; font-weight: 900; color: #D6B864 !important; margin-top: 6px; }

.sh {
    font-size: 1.02rem; font-weight: 800;
    border-bottom: 2px solid #0D2B40; padding-bottom: 5px; margin: 14px 0 10px;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
# HELPERS / ASSETS
# ══════════════════════════════════════════════════════════════════════
def img_to_b64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return ""

SCRIPT_DIR = Path(__file__).parent
LOGOS_DIR    = SCRIPT_DIR / "logos"
APOSTAS_DIR  = SCRIPT_DIR / "apostas"
GABARITO_DIR = SCRIPT_DIR / "gabarito"
CONSOLIDADA_PATH = APOSTAS_DIR / "Bolao_Copa2026_DaGalera - Consolidada.xlsx"

def find_asset(name):
    for p in [LOGOS_DIR / name, SCRIPT_DIR / name]:
        if p.exists():
            return str(p)
    return ""

# Apenas estas 3 imagens são usadas nesta versão.
MASCOTES   = find_asset("mascotes.png")
FRONT_PAGE = find_asset("front_page.png")
FAVICON    = find_asset("favicon.png")

MOSTRAR_SIMULACAO = True   # True → mostra aba 🔮 Simulação

JOGADORES_ARTILHEIRO: list[tuple[str, str]] = [
    ('Mohammed Al-Owais', 'Arabia Saudita'),
    ('Nawaf Al-Aqidi', 'Arabia Saudita'),
    ('Ahmed Al-Kassar', 'Arabia Saudita'),
    ('Ali Al-Bulayhi', 'Arabia Saudita'),
    ('Hassan Kadesh', 'Arabia Saudita'),
    ('Yasir Al-Shahrani', 'Arabia Saudita'),
    ('Nawaf Boushal', 'Arabia Saudita'),
    ('Sultan Al-Ghannam', 'Arabia Saudita'),
    ('Nasser Al-Dawsari', 'Arabia Saudita'),
    ('Musab Al-Juwayr', 'Arabia Saudita'),
    ('Faisal Al-Ghamdi', 'Arabia Saudita'),
    ('Ayman Yahya', 'Arabia Saudita'),
    ('Abdulrahman Ghareeb', 'Arabia Saudita'),
    ('Firas Al-Buraikan', 'Arabia Saudita'),
    ('Abdullah Al-Hamdan', 'Arabia Saudita'),
    ('Saleh Al-Shehri', 'Arabia Saudita'),
    ('Saud Abdulhamid', 'Arabia Saudita'),
    ('Mohammed Kanno', 'Arabia Saudita'),
    ('Salem Al-Dawsari', 'Arabia Saudita'),
    ('Ahmed Al-Ghamdi', 'Arabia Saudita'),
    ('Mat Ryan', 'Australia'),
    ('Joe Gauci', 'Australia'),
    ('Patrick Beach', 'Australia'),
    ('Harry Souttar', 'Australia'),
    ('Kye Rowles', 'Australia'),
    ('Alessandro Circati', 'Australia'),
    ('Cameron Burgess', 'Australia'),
    ('Aziz Behich', 'Australia'),
    ('Jordan Bos', 'Australia'),
    ('Nathaniel Atkinson', 'Australia'),
    ('Jackson Irvine', 'Australia'),
    ("Aiden O'Neill", 'Australia'),
    ('Riley McGree', 'Australia'),
    ('Connor Metcalfe', 'Australia'),
    ('Martin Boyle', 'Australia'),
    ('Craig Goodwin', 'Australia'),
    ('Kusini Yengi', 'Australia'),
    ('Brandon Borrello', 'Australia'),
    ('Daniel Bennie', 'Australia'),
    ('Raphael Borges Rodrigues', 'Australia'),
    ('Awer Mabil', 'Australia'),
    ('Jalal Hassan', 'Iraque'),
    ('Ahmed Basil', 'Iraque'),
    ('Kumel Al-Rkabe', 'Iraque'),
    ('Ahmed Yahya', 'Iraque'),
    ('Rebin Sulaka', 'Iraque'),
    ('Frans Putros', 'Iraque'),
    ('Hussein Ali', 'Iraque'),
    ('Zaid Tahseen', 'Iraque'),
    ('Ali Jasim', 'Iraque'),
    ('Ibrahim Bayesh', 'Iraque'),
    ('Zidane Iqbal', 'Iraque'),
    ('Amir Al-Ammari', 'Iraque'),
    ('Osama Rashid', 'Iraque'),
    ('Youssef Amyn', 'Iraque'),
    ('Mohanad Ali', 'Iraque'),
    ('Aymen Hussein', 'Iraque'),
    ('Ali Al-Hamadi', 'Iraque'),
    ('Marko Farji', 'Iraque'),
    ('Peter Gwargis', 'Iraque'),
    ('Hassan Abdulkarim', 'Iraque'),
    ('Alireza Beiranvand', 'Ira'),
    ('Payam Niazmand', 'Ira'),
    ('Hossein Hosseini', 'Ira'),
    ('Hossein Kanaanizadegan', 'Ira'),
    ('Shoja Khalilzadeh', 'Ira'),
    ('Ramin Rezaeian', 'Ira'),
    ("Ali Ne'mati", 'Ira'),
    ('Milad Mohammadi', 'Ira'),
    ('Roozbeh Cheshmi', 'Ira'),
    ('Saman Ghoddos', 'Ira'),
    ('Saeid Ezatolahi', 'Ira'),
    ('Alireza Jahanbakhsh', 'Ira'),
    ('Mehdi Ghayedi', 'Ira'),
    ('Mehdi Taremi', 'Ira'),
    ('Mohammad Mohebi', 'Ira'),
    ('Mohammad Ghorbani', 'Ira'),
    ('Sardar Azmoun', 'Ira'),
    ('Allahyar Sayyadmanesh', 'Ira'),
    ('Omid Noorafkan', 'Ira'),
    ('Mehdi Torabi', 'Ira'),
    ('Ehsan Hajsafi', 'Ira'),
    ('Zion Suzuki', 'Japao'),
    ('Keisuke Osako', 'Japao'),
    ('Tomoki Hayakawa', 'Japao'),
    ('Ko Itakura', 'Japao'),
    ('Hiroki Ito', 'Japao'),
    ('Shogo Taniguchi', 'Japao'),
    ('Ayumu Seko', 'Japao'),
    ('Yukinari Sugawara', 'Japao'),
    ('Wataru Endo', 'Japao'),
    ('Hidemasa Morita', 'Japao'),
    ('Reo Hatate', 'Japao'),
    ('Takefusa Kubo', 'Japao'),
    ('Kaoru Mitoma', 'Japao'),
    ('Ritsu Doan', 'Japao'),
    ('Daichi Kamada', 'Japao'),
    ('Junya Ito', 'Japao'),
    ('Ayase Ueda', 'Japao'),
    ('Shuto Machino', 'Japao'),
    ('Junnosuke Suzuki', 'Japao'),
    ('Joel Chima Fujita', 'Japao'),
    ('Takehiro Tomiyasu', 'Japao'),
    ('Yazeed Abu Laila', 'Jordania'),
    ('Abdullah Nasib', 'Jordania'),
    ('Yazan Al-Arab', 'Jordania'),
    ('Noor Al-Rawabdeh', 'Jordania'),
    ('Nizar Al-Rashdan', 'Jordania'),
    ('Ali Olwan', 'Jordania'),
    ('Mousa Al-Tamari', 'Jordania'),
    ('Yazan Al-Naimat', 'Jordania'),
    ('Mahmoud Al-Mardi', 'Jordania'),
    ('Saleh Ratib', 'Jordania'),
    ('Mohammad Abu Hasheesh', 'Jordania'),
    ('Anas Bani Yaseen', 'Jordania'),
    ('Amer Jamous', 'Jordania'),
    ('Ibrahim Sabra', 'Jordania'),
    ('Meshaal Barsham', 'Qatar'),
    ('Mahmoud Abunada', 'Qatar'),
    ('Salah Zakaria', 'Qatar'),
    ('Tarek Salman', 'Qatar'),
    ('Pedro Miguel', 'Qatar'),
    ('Boualem Khoukhi', 'Qatar'),
    ('Lucas Mendes', 'Qatar'),
    ('Homam Al-Amin', 'Qatar'),
    ('Karim Boudiaf', 'Qatar'),
    ('Assim Madibo', 'Qatar'),
    ('Mohammed Waad', 'Qatar'),
    ('Ahmed Alaaeldin', 'Qatar'),
    ('Akram Afif', 'Qatar'),
    ('Almoez Ali', 'Qatar'),
    ('Mohammed Manai', 'Qatar'),
    ('Mustapha Tariq', 'Qatar'),
    ('Tahsin Mohammed Jamshid', 'Qatar'),
    ('Sebastian Soria', 'Qatar'),
    ('Jo Hyeon-woo', 'Coreia do Sul'),
    ('Kim Seung-gyu', 'Coreia do Sul'),
    ('Kim Min-jae', 'Coreia do Sul'),
    ('Kim Jin-su', 'Coreia do Sul'),
    ('Seol Young-woo', 'Coreia do Sul'),
    ('Kim Young-gwon', 'Coreia do Sul'),
    ('Hwang In-beom', 'Coreia do Sul'),
    ('Lee Jae-sung', 'Coreia do Sul'),
    ('Lee Kang-in', 'Coreia do Sul'),
    ('Hwang Hee-chan', 'Coreia do Sul'),
    ('Son Heung-min', 'Coreia do Sul'),
    ('Cho Gue-sung', 'Coreia do Sul'),
    ('Oh Hyeon-gyu', 'Coreia do Sul'),
    ('Bae Jun-ho', 'Coreia do Sul'),
    ('Yang Min-hyeok', 'Coreia do Sul'),
    ('Jeong Sang-bin', 'Coreia do Sul'),
    ('Hong Hyun-seok', 'Coreia do Sul'),
    ('Utkir Yusupov', 'Uzbequistao'),
    ('Abduvohid Nematov', 'Uzbequistao'),
    ('Abdukodir Khusanov', 'Uzbequistao'),
    ('Rustam Ashurmatov', 'Uzbequistao'),
    ('Farrukh Sayfiev', 'Uzbequistao'),
    ('Odiljon Hamrobekov', 'Uzbequistao'),
    ('Abbosbek Fayzullaev', 'Uzbequistao'),
    ('Jasurbek Jaloliddinov', 'Uzbequistao'),
    ('Jaloliddin Masharipov', 'Uzbequistao'),
    ('Oston Urunov', 'Uzbequistao'),
    ('Eldor Shomurodov', 'Uzbequistao'),
    ('Umarali Rakhmonaliev', 'Uzbequistao'),
    ("Muhammadali O'rinboyev", 'Uzbequistao'),
    ('Sardor Rashidov', 'Uzbequistao'),
    ('Bobur Abdukholikov', 'Uzbequistao'),
    ('Anthony Mandrea', 'Argelia'),
    ('Alexandre Oukidja', 'Argelia'),
    ('Rayan Aït-Nouri', 'Argelia'),
    ('Ramy Bensebaini', 'Argelia'),
    ('Aïssa Mandi', 'Argelia'),
    ('Zineddine Belaïd', 'Argelia'),
    ('Ismaël Bennacer', 'Argelia'),
    ('Houssem Aouar', 'Argelia'),
    ('Farès Chaïbi', 'Argelia'),
    ('Riyad Mahrez', 'Argelia'),
    ('Mohamed Amoura', 'Argelia'),
    ('Amine Gouiri', 'Argelia'),
    ('Baghdad Bounedjah', 'Argelia'),
    ('Saïd Benrahma', 'Argelia'),
    ('Anis Hadj Moussa', 'Argelia'),
    ('Ibrahim Maza', 'Argelia'),
    ('Bruno Varela', 'Cabo Verde'),
    ('Vozinha', 'Cabo Verde'),
    ('Roberto Lopes', 'Cabo Verde'),
    ('Logan Costa', 'Cabo Verde'),
    ('Stopira', 'Cabo Verde'),
    ('Jojo', 'Cabo Verde'),
    ('Kevin Pina', 'Cabo Verde'),
    ('Laros Duarte', 'Cabo Verde'),
    ('Jamiro Monteiro', 'Cabo Verde'),
    ('Bebé', 'Cabo Verde'),
    ('Ryan Mendes', 'Cabo Verde'),
    ('Jovane Cabral', 'Cabo Verde'),
    ('Benchimol', 'Cabo Verde'),
    ('Gilson Tavares', 'Cabo Verde'),
    ('Kevin Lenini', 'Cabo Verde'),
    ('Yahia Fofana', 'Costa do Marfim'),
    ('Alban Lafont', 'Costa do Marfim'),
    ('Odilon Kossounou', 'Costa do Marfim'),
    ('Evan Ndicka', 'Costa do Marfim'),
    ('Wilfried Singo', 'Costa do Marfim'),
    ('Ghislain Konan', 'Costa do Marfim'),
    ('Franck Kessie', 'Costa do Marfim'),
    ('Seko Fofana', 'Costa do Marfim'),
    ('Ibrahim Sangare', 'Costa do Marfim'),
    ('Amad Diallo', 'Costa do Marfim'),
    ('Simon Adingra', 'Costa do Marfim'),
    ('Sebastien Haller', 'Costa do Marfim'),
    ('Nicolas Pepe', 'Costa do Marfim'),
    ('Oumar Diakite', 'Costa do Marfim'),
    ('Yan Diomande', 'Costa do Marfim'),
    ('Ange-Yoan Bonny', 'Costa do Marfim'),
    ('Karim Konaté', 'Costa do Marfim'),
    ('Lionel Mpasi', 'RD Congo'),
    ('Theo Fayulu', 'RD Congo'),
    ('Matthieu Epolo', 'RD Congo'),
    ('Chancel Mbemba', 'RD Congo'),
    ('Aaron Wan-Bissaka', 'RD Congo'),
    ('Arthur Masuaku', 'RD Congo'),
    ('Joris Kayembe', 'RD Congo'),
    ('Axel Tuanzebe', 'RD Congo'),
    ('Dylan Batubinsika', 'RD Congo'),
    ('Edo Kayembe', 'RD Congo'),
    ('Samuel Moutoussamy', 'RD Congo'),
    ('Noah Sadiki', 'RD Congo'),
    ("Ngal'ayel Mukau", 'RD Congo'),
    ('Yoane Wissa', 'RD Congo'),
    ('Cedric Bakambu', 'RD Congo'),
    ('Simon Banza', 'RD Congo'),
    ('Theo Bongonda', 'RD Congo'),
    ('Mohamed El Shenawy', 'Egito'),
    ('Mohamed Sobhy', 'Egito'),
    ('Mohamed Abdelmonem', 'Egito'),
    ('Ahmed Hegazi', 'Egito'),
    ('Omar Kamal', 'Egito'),
    ('Mohamed Hany', 'Egito'),
    ('Hamdi Fathi', 'Egito'),
    ('Marwan Atiya', 'Egito'),
    ('Mohamed Elneny', 'Egito'),
    ('Trezeguet', 'Egito'),
    ('Omar Marmoush', 'Egito'),
    ('Mohamed Salah', 'Egito'),
    ('Mostafa Mohamed', 'Egito'),
    ('Ibrahim Adel', 'Egito'),
    ('Mustafa Fathi', 'Egito'),
    ('Lawrence Ati-Zigi', 'Gana'),
    ('Joseph Wollacott', 'Gana'),
    ('Alexander Djiku', 'Gana'),
    ('Mohammed Salisu', 'Gana'),
    ('Tariq Lamptey', 'Gana'),
    ('Gideon Mensah', 'Gana'),
    ('Thomas Partey', 'Gana'),
    ('Elisha Owusu', 'Gana'),
    ('Mohammed Kudus', 'Gana'),
    ('Abdul Fatawu Issahaku', 'Gana'),
    ('Antoine Semenyo', 'Gana'),
    ('Inaki Williams', 'Gana'),
    ('Jordan Ayew', 'Gana'),
    ('Ernest Nuamah', 'Gana'),
    ('Ibrahim Osman', 'Gana'),
    ('Salis Abdul Samed', 'Gana'),
    ('Bono', 'Marrocos'),
    ('Munir El Kajoui', 'Marrocos'),
    ('Achraf Hakimi', 'Marrocos'),
    ('Noussair Mazraoui', 'Marrocos'),
    ('Nayef Aguerd', 'Marrocos'),
    ('Romain Saiss', 'Marrocos'),
    ('Sofyan Amrabat', 'Marrocos'),
    ('Azzedine Ounahi', 'Marrocos'),
    ('Bilal El Khannouss', 'Marrocos'),
    ('Brahim Diaz', 'Marrocos'),
    ('Hakim Ziyech', 'Marrocos'),
    ('Amine Adli', 'Marrocos'),
    ('Youssef En-Nesyri', 'Marrocos'),
    ('Ayoub El Kaabi', 'Marrocos'),
    ('Eliesse Ben Seghir', 'Marrocos'),
    ('Chadi Riad', 'Marrocos'),
    ('Sofiane Boufal', 'Marrocos'),
    ('Abde Ezzalzouli', 'Marrocos'),
    ('Edouard Mendy', 'Senegal'),
    ('Mory Diaw', 'Senegal'),
    ('Kalidou Koulibaly', 'Senegal'),
    ('Moussa Niakhate', 'Senegal'),
    ('Ismail Jakobs', 'Senegal'),
    ('Krépin Diatta', 'Senegal'),
    ('Idrissa Gana Gueye', 'Senegal'),
    ('Pape Matar Sarr', 'Senegal'),
    ('Lamine Camara', 'Senegal'),
    ('Nicolas Jackson', 'Senegal'),
    ('Ismaila Sarr', 'Senegal'),
    ('Iliman Ndiaye', 'Senegal'),
    ('Boulaye Dia', 'Senegal'),
    ('Habib Diallo', 'Senegal'),
    ('Sadio Mane', 'Senegal'),
    ('Mikayil Faye', 'Senegal'),
    ('Abdallah Sima', 'Senegal'),
    ('Ronwen Williams', 'Africa do Sul'),
    ('Siyabonga Ngezana', 'Africa do Sul'),
    ('Grant Kekana', 'Africa do Sul'),
    ('Khuliso Mudau', 'Africa do Sul'),
    ('Teboho Mokoena', 'Africa do Sul'),
    ('Mothobi Mvala', 'Africa do Sul'),
    ('Themba Zwane', 'Africa do Sul'),
    ('Percy Tau', 'Africa do Sul'),
    ('Oswin Appollis', 'Africa do Sul'),
    ('Evidence Makgopa', 'Africa do Sul'),
    ('Iqraam Rayners', 'Africa do Sul'),
    ('Relebohile Mofokeng', 'Africa do Sul'),
    ('Shandre Campbell', 'Africa do Sul'),
    ('Aubrey Modiba', 'Africa do Sul'),
    ('Nkosinathi Sibisi', 'Africa do Sul'),
    ('Aymen Dahmen', 'Tunisia'),
    ('Bechir Ben Said', 'Tunisia'),
    ('Montassar Talbi', 'Tunisia'),
    ('Dylan Bronn', 'Tunisia'),
    ('Ali Abdi', 'Tunisia'),
    ('Wajdi Kechrida', 'Tunisia'),
    ('Ellyes Skhiri', 'Tunisia'),
    ('Aissa Laidouni', 'Tunisia'),
    ('Hannibal Mejbri', 'Tunisia'),
    ('Mohamed Ali Ben Romdhane', 'Tunisia'),
    ('Anis Ben Slimane', 'Tunisia'),
    ('Elias Achouri', 'Tunisia'),
    ('Seifeddine Jaziri', 'Tunisia'),
    ('Haythem Jouini', 'Tunisia'),
    ('Youssef Msakni', 'Tunisia'),
    ('Maxime Crépeau', 'Canada'),
    ('Dayne St. Clair', 'Canada'),
    ('Alistair Johnston', 'Canada'),
    ('Alphonso Davies', 'Canada'),
    ('Moïse Bombito', 'Canada'),
    ('Derek Cornelius', 'Canada'),
    ('Ismaël Koné', 'Canada'),
    ('Stephen Eustaquio', 'Canada'),
    ('Jonathan David', 'Canada'),
    ('Cyle Larin', 'Canada'),
    ('Tajon Buchanan', 'Canada'),
    ('Jacob Shaffelburg', 'Canada'),
    ('Mathieu Choinière', 'Canada'),
    ('Richie Laryea', 'Canada'),
    ('Luc de Fougerolles', 'Canada'),
    ('Ali Ahmed', 'Canada'),
    ('Eloy Room', 'Curacao'),
    ('Cuco Martina', 'Curacao'),
    ('Jurien Gaari', 'Curacao'),
    ('Sherel Floranus', 'Curacao'),
    ('Juninho Bacuna', 'Curacao'),
    ('Leandro Bacuna', 'Curacao'),
    ('Vurnon Anita', 'Curacao'),
    ('Kenji Gorré', 'Curacao'),
    ('Jeremy Antonisse', 'Curacao'),
    ('Gervane Kastaneer', 'Curacao'),
    ('Rangelo Janga', 'Curacao'),
    ('Joshua Zimmerman', 'Curacao'),
    ('Livano Comenencia', 'Curacao'),
    ('Johny Placide', 'Haiti'),
    ('Alexandre Pierre', 'Haiti'),
    ('Ricardo Adé', 'Haiti'),
    ('Carlens Arcus', 'Haiti'),
    ('Alex Christian Jr.', 'Haiti'),
    ('Danley Jean Jacques', 'Haiti'),
    ('Bryan Alceus', 'Haiti'),
    ('Leverton Pierre', 'Haiti'),
    ('Louicius Don Deedson', 'Haiti'),
    ('Duckens Nazon', 'Haiti'),
    ('Frantzdy Pierrot', 'Haiti'),
    ('Mondy Prunier', 'Haiti'),
    ('Ruben Providence', 'Haiti'),
    ('Fafa Picault', 'Haiti'),
    ('Luis Ángel Malagón', 'Mexico'),
    ('Guillermo Ochoa', 'Mexico'),
    ('Johan Vásquez', 'Mexico'),
    ('César Montes', 'Mexico'),
    ('Jesús Gallardo', 'Mexico'),
    ('Jorge Sánchez', 'Mexico'),
    ('Edson Álvarez', 'Mexico'),
    ('Luis Chávez', 'Mexico'),
    ('Orbelín Pineda', 'Mexico'),
    ('Marcel Ruiz', 'Mexico'),
    ('Roberto Alvarado', 'Mexico'),
    ('César Huerta', 'Mexico'),
    ('Santiago Giménez', 'Mexico'),
    ('Raúl Jiménez', 'Mexico'),
    ('Hirving Lozano', 'Mexico'),
    ('Julián Quiñones', 'Mexico'),
    ('Fidel Ambriz', 'Mexico'),
    ('Gilberto Mora', 'Mexico'),
    ('Orlando Mosquera', 'Panama'),
    ('César Blackman', 'Panama'),
    ('Éric Davis', 'Panama'),
    ('Fidel Escobar', 'Panama'),
    ('Andrés Andrade', 'Panama'),
    ('José Córdoba', 'Panama'),
    ('Michael Murillo', 'Panama'),
    ('Adalberto Carrasquilla', 'Panama'),
    ('Aníbal Godoy', 'Panama'),
    ('Cristian Martínez', 'Panama'),
    ('Ismael Díaz', 'Panama'),
    ('José Fajardo', 'Panama'),
    ('Eduardo Guerrero', 'Panama'),
    ('José Luis Rodríguez', 'Panama'),
    ('Kahiser Lenis', 'Panama'),
    ('Matt Turner', 'Estados Unidos'),
    ('Zack Steffen', 'Estados Unidos'),
    ('Joe Scally', 'Estados Unidos'),
    ('Sergiño Dest', 'Estados Unidos'),
    ('Antonee Robinson', 'Estados Unidos'),
    ('Chris Richards', 'Estados Unidos'),
    ('Tim Ream', 'Estados Unidos'),
    ('Tyler Adams', 'Estados Unidos'),
    ('Weston McKennie', 'Estados Unidos'),
    ('Yunus Musah', 'Estados Unidos'),
    ('Gio Reyna', 'Estados Unidos'),
    ('Christian Pulisic', 'Estados Unidos'),
    ('Tim Weah', 'Estados Unidos'),
    ('Folarin Balogun', 'Estados Unidos'),
    ('Ricardo Pepi', 'Estados Unidos'),
    ('Josh Sargent', 'Estados Unidos'),
    ('Malik Tillman', 'Estados Unidos'),
    ('Diego Luna', 'Estados Unidos'),
    ('Paxten Aaronson', 'Estados Unidos'),
    ('Emiliano Martínez', 'Argentina'),
    ('Gerónimo Rulli', 'Argentina'),
    ('Cristian Romero', 'Argentina'),
    ('Lisandro Martínez', 'Argentina'),
    ('Nicolás Otamendi', 'Argentina'),
    ('Nahuel Molina', 'Argentina'),
    ('Marcos Acuña', 'Argentina'),
    ('Alexis Mac Allister', 'Argentina'),
    ('Enzo Fernández', 'Argentina'),
    ('Rodrigo De Paul', 'Argentina'),
    ('Leandro Paredes', 'Argentina'),
    ('Exequiel Palacios', 'Argentina'),
    ('Lionel Messi', 'Argentina'),
    ('Julián Álvarez', 'Argentina'),
    ('Lautaro Martínez', 'Argentina'),
    ('Nico González', 'Argentina'),
    ('Thiago Almada', 'Argentina'),
    ('Alejandro Garnacho', 'Argentina'),
    ('Alisson', 'Brasil'),
    ('Ederson', 'Brasil'),
    ('Marquinhos', 'Brasil'),
    ('Gabriel Magalhães', 'Brasil'),
    ('Éder Militão', 'Brasil'),
    ('Beraldo', 'Brasil'),
    ('Guilherme Arana', 'Brasil'),
    ('Danilo', 'Brasil'),
    ('Danilo Santos', 'Brasil'),
    ('Bruno Guimarães', 'Brasil'),
    ('João Gomes', 'Brasil'),
    ('Lucas Paquetá', 'Brasil'),
    ('Rodrygo', 'Brasil'),
    ('Vinícius Júnior', 'Brasil'),
    ('Raphinha', 'Brasil'),
    ('Endrick', 'Brasil'),
    ('Gabriel Martinelli', 'Brasil'),
    ('Savinho', 'Brasil'),
    ('João Pedro', 'Brasil'),
    ('Estevão', 'Brasil'),
    ('Neymar', 'Brasil'),
    ('Camilo Vargas', 'Colombia'),
    ('Álvaro Montero', 'Colombia'),
    ('Davinson Sánchez', 'Colombia'),
    ('Jhon Lucumí', 'Colombia'),
    ('Daniel Muñoz', 'Colombia'),
    ('Deiver Machado', 'Colombia'),
    ('Johan Mojica', 'Colombia'),
    ('Jefferson Lerma', 'Colombia'),
    ('Richard Ríos', 'Colombia'),
    ('Kevin Castaño', 'Colombia'),
    ('James Rodríguez', 'Colombia'),
    ('Jhon Arias', 'Colombia'),
    ('Luis Díaz', 'Colombia'),
    ('Jhon Durán', 'Colombia'),
    ('Rafael Santos Borré', 'Colombia'),
    ('Andrés Gómez', 'Colombia'),
    ('Luis Sinisterra', 'Colombia'),
    ('Yáser Asprilla', 'Colombia'),
    ('Hernán Galíndez', 'Equador'),
    ('Alexander Domínguez', 'Equador'),
    ('Willian Pacho', 'Equador'),
    ('Piero Hincapié', 'Equador'),
    ('Félix Torres', 'Equador'),
    ('Angelo Preciado', 'Equador'),
    ('Moisés Caicedo', 'Equador'),
    ('Alan Franco', 'Equador'),
    ('Kendry Páez', 'Equador'),
    ('Jeremy Sarmiento', 'Equador'),
    ('Gonzalo Plata', 'Equador'),
    ('Kevin Rodríguez', 'Equador'),
    ('Enner Valencia', 'Equador'),
    ('John Yeboah', 'Equador'),
    ('Nilson Angulo', 'Equador'),
    ('Justin Lerma', 'Equador'),
    ('Carlos Coronel', 'Paraguai'),
    ('Gatito Fernández', 'Paraguai'),
    ('Fabián Balbuena', 'Paraguai'),
    ('Omar Alderete', 'Paraguai'),
    ('Juan José Cáceres', 'Paraguai'),
    ('Diego Gómez', 'Paraguai'),
    ('Mathías Villasanti', 'Paraguai'),
    ('Andrés Cubas', 'Paraguai'),
    ('Miguel Almirón', 'Paraguai'),
    ('Julio Enciso', 'Paraguai'),
    ('Antonio Sanabria', 'Paraguai'),
    ('Ramón Sosa', 'Paraguai'),
    ('Enso González', 'Paraguai'),
    ('Diego León', 'Paraguai'),
    ('Gabriel Ávalos', 'Paraguai'),
    ('Sergio Rochet', 'Uruguai'),
    ('Santiago Mele', 'Uruguai'),
    ('Ronald Araújo', 'Uruguai'),
    ('José María Giménez', 'Uruguai'),
    ('Sebastián Cáceres', 'Uruguai'),
    ('Mathías Olivera', 'Uruguai'),
    ('Nahitan Nández', 'Uruguai'),
    ('Federico Valverde', 'Uruguai'),
    ('Manuel Ugarte', 'Uruguai'),
    ('Rodrigo Bentancur', 'Uruguai'),
    ('Nicolás de la Cruz', 'Uruguai'),
    ('Darwin Núñez', 'Uruguai'),
    ('Facundo Pellistri', 'Uruguai'),
    ('Maximiliano Araújo', 'Uruguai'),
    ('Luciano Rodríguez', 'Uruguai'),
    ('Puma Rodríguez', 'Uruguai'),
    ('Giorgian de Arrascaeta', 'Uruguai'),
    ('Franco González', 'Uruguai'),
    ('Brian Rodríguez', 'Uruguai'),
    ('Max Crocombe', 'Nova Zelandia'),
    ('Michael Boxall', 'Nova Zelandia'),
    ('Tommy Smith', 'Nova Zelandia'),
    ('Liberato Cacace', 'Nova Zelandia'),
    ('Tim Payne', 'Nova Zelandia'),
    ('Tyler Bindon', 'Nova Zelandia'),
    ('Marko Stamenic', 'Nova Zelandia'),
    ('Joe Bell', 'Nova Zelandia'),
    ('Sarpreet Singh', 'Nova Zelandia'),
    ('Matt Garbett', 'Nova Zelandia'),
    ('Elijah Just', 'Nova Zelandia'),
    ('Ben Waine', 'Nova Zelandia'),
    ('Chris Wood', 'Nova Zelandia'),
    ('Callum McCowatt', 'Nova Zelandia'),
    ('Alexander Schlager', 'Austria'),
    ('Patrick Pentz', 'Austria'),
    ('David Alaba', 'Austria'),
    ('Kevin Danso', 'Austria'),
    ('Philipp Lienhart', 'Austria'),
    ('Stefan Posch', 'Austria'),
    ('Alexander Prass', 'Austria'),
    ('Konrad Laimer', 'Austria'),
    ('Nicolas Seiwald', 'Austria'),
    ('Christoph Baumgartner', 'Austria'),
    ('Marcel Sabitzer', 'Austria'),
    ('Romano Schmid', 'Austria'),
    ('Marko Arnautović', 'Austria'),
    ('Michael Gregoritsch', 'Austria'),
    ('Samson Baidoo', 'Austria'),
    ('Raul Florucz', 'Austria'),
    ('Koen Casteels', 'Belgica'),
    ('Matz Sels', 'Belgica'),
    ('Wout Faes', 'Belgica'),
    ('Arthur Theate', 'Belgica'),
    ('Zeno Debast', 'Belgica'),
    ('Timothy Castagne', 'Belgica'),
    ('Amadou Onana', 'Belgica'),
    ('Orel Mangala', 'Belgica'),
    ('Youri Tielemans', 'Belgica'),
    ('Kevin De Bruyne', 'Belgica'),
    ('Jérémy Doku', 'Belgica'),
    ('Charles De Ketelaere', 'Belgica'),
    ('Lois Openda', 'Belgica'),
    ('Romelu Lukaku', 'Belgica'),
    ('Arthur Vermeeren', 'Belgica'),
    ('Malick Fofana', 'Belgica'),
    ('Julien Duranville', 'Belgica'),
    ('Leandro Trossard', 'Belgica'),
    ('Nikola Vasilj', 'Bosnia e Herzegovina'),
    ('Ibrahim Šehić', 'Bosnia e Herzegovina'),
    ('Amar Dedić', 'Bosnia e Herzegovina'),
    ('Sead Kolašinac', 'Bosnia e Herzegovina'),
    ('Dennis Hadžikadunić', 'Bosnia e Herzegovina'),
    ('Jusuf Gazibegović', 'Bosnia e Herzegovina'),
    ('Benjamin Tahirović', 'Bosnia e Herzegovina'),
    ('Haris Hajradinović', 'Bosnia e Herzegovina'),
    ('Ivan Bašić', 'Bosnia e Herzegovina'),
    ('Dženis Burnić', 'Bosnia e Herzegovina'),
    ('Esmir Bajraktarević', 'Bosnia e Herzegovina'),
    ('Edin Džeko', 'Bosnia e Herzegovina'),
    ('Ermedin Demirović', 'Bosnia e Herzegovina'),
    ('Smail Prevljak', 'Bosnia e Herzegovina'),
    ('Dominik Livaković', 'Croacia'),
    ('Josip Šutalo', 'Croacia'),
    ('Joško Gvardiol', 'Croacia'),
    ('Josip Stanišić', 'Croacia'),
    ('Borna Sosa', 'Croacia'),
    ('Luka Modrić', 'Croacia'),
    ('Mateo Kovačić', 'Croacia'),
    ('Marcelo Brozović', 'Croacia'),
    ('Luka Sučić', 'Croacia'),
    ('Martin Baturina', 'Croacia'),
    ('Mario Pašalić', 'Croacia'),
    ('Andrej Kramarić', 'Croacia'),
    ('Ante Budimir', 'Croacia'),
    ('Ivan Perišić', 'Croacia'),
    ('Lovro Majer', 'Croacia'),
    ('Petar Sučić', 'Croacia'),
    ('Dion Drena Beljo', 'Croacia'),
    ('Jindřich Staněk', 'Tchequia'),
    ('Tomáš Vaclík', 'Tchequia'),
    ('Tomáš Holeš', 'Tchequia'),
    ('Robin Hranáč', 'Tchequia'),
    ('Vladimír Coufal', 'Tchequia'),
    ('David Douděra', 'Tchequia'),
    ('Tomáš Souček', 'Tchequia'),
    ('Antonín Barák', 'Tchequia'),
    ('Lukáš Provod', 'Tchequia'),
    ('Adam Hložek', 'Tchequia'),
    ('Patrik Schick', 'Tchequia'),
    ('Václav Černý', 'Tchequia'),
    ('Jan Kuchta', 'Tchequia'),
    ('Matěj Jurásek', 'Tchequia'),
    ('Adam Karabec', 'Tchequia'),
    ('Jordan Pickford', 'Inglaterra'),
    ('Aaron Ramsdale', 'Inglaterra'),
    ('John Stones', 'Inglaterra'),
    ('Marc Guéhi', 'Inglaterra'),
    ('Ezri Konsa', 'Inglaterra'),
    ('Luke Shaw', 'Inglaterra'),
    ('Trent Alexander-Arnold', 'Inglaterra'),
    ('Rico Lewis', 'Inglaterra'),
    ('Declan Rice', 'Inglaterra'),
    ('Jude Bellingham', 'Inglaterra'),
    ('Cole Palmer', 'Inglaterra'),
    ('Bukayo Saka', 'Inglaterra'),
    ('Phil Foden', 'Inglaterra'),
    ('Anthony Gordon', 'Inglaterra'),
    ('Harry Kane', 'Inglaterra'),
    ('Ollie Watkins', 'Inglaterra'),
    ('Ivan Toney', 'Inglaterra'),
    ('Jarell Quansah', 'Inglaterra'),
    ('Ethan Nwaneri', 'Inglaterra'),
    ('Reece James', 'Inglaterra'),
    ('Mike Maignan', 'Franca'),
    ('Brice Samba', 'Franca'),
    ('William Saliba', 'Franca'),
    ('Ibrahima Konaté', 'Franca'),
    ('Dayot Upamecano', 'Franca'),
    ('Jules Koundé', 'Franca'),
    ('Theo Hernandez', 'Franca'),
    ('Aurélien Tchouaméni', 'Franca'),
    ('Eduardo Camavinga', 'Franca'),
    ('Warren Zaïre-Emery', 'Franca'),
    ('Michael Olise', 'Franca'),
    ('Ousmane Dembélé', 'Franca'),
    ('Bradley Barcola', 'Franca'),
    ('Kylian Mbappé', 'Franca'),
    ('Marcus Thuram', 'Franca'),
    ('Randal Kolo Muani', 'Franca'),
    ('Désiré Doué', 'Franca'),
    ('Mathys Tel', 'Franca'),
    ('Youssouf Fofana', 'Franca'),
    ('Marc-André ter Stegen', 'Alemanha'),
    ('Manuel Neuer', 'Alemanha'),
    ('Antonio Rüdiger', 'Alemanha'),
    ('Jonathan Tah', 'Alemanha'),
    ('Nico Schlotterbeck', 'Alemanha'),
    ('David Raum', 'Alemanha'),
    ('Joshua Kimmich', 'Alemanha'),
    ('Ilkay Gündoğan', 'Alemanha'),
    ('Robert Andrich', 'Alemanha'),
    ('Aleksandar Pavlović', 'Alemanha'),
    ('Florian Wirtz', 'Alemanha'),
    ('Jamal Musiala', 'Alemanha'),
    ('Kai Havertz', 'Alemanha'),
    ('Leroy Sané', 'Alemanha'),
    ('Karim Adeyemi', 'Alemanha'),
    ('Niclas Füllkrug', 'Alemanha'),
    ('Maximilian Beier', 'Alemanha'),
    ('Brajan Gruda', 'Alemanha'),
    ('Bart Verbruggen', 'Holanda'),
    ('Mark Flekken', 'Holanda'),
    ('Virgil van Dijk', 'Holanda'),
    ('Matthijs de Ligt', 'Holanda'),
    ('Nathan Aké', 'Holanda'),
    ('Jeremie Frimpong', 'Holanda'),
    ('Denzel Dumfries', 'Holanda'),
    ('Frenkie de Jong', 'Holanda'),
    ('Teun Koopmeiners', 'Holanda'),
    ('Tijjani Reijnders', 'Holanda'),
    ('Xavi Simons', 'Holanda'),
    ('Cody Gakpo', 'Holanda'),
    ('Joshua Zirkzee', 'Holanda'),
    ('Brian Brobbey', 'Holanda'),
    ('Noa Lang', 'Holanda'),
    ('Jorrel Hato', 'Holanda'),
    ('Ian Maatsen', 'Holanda'),
    ('Memphis Depay', 'Holanda'),
    ('Ørjan Nyland', 'Noruega'),
    ('Kristoffer Ajer', 'Noruega'),
    ('Leo Østigård', 'Noruega'),
    ('Julian Ryerson', 'Noruega'),
    ('Andreas Hanche-Olsen', 'Noruega'),
    ('Martin Ødegaard', 'Noruega'),
    ('Sander Berge', 'Noruega'),
    ('Patrick Berg', 'Noruega'),
    ('Oscar Bobb', 'Noruega'),
    ('Antonio Nusa', 'Noruega'),
    ('Erling Haaland', 'Noruega'),
    ('Alexander Sørloth', 'Noruega'),
    ('Andreas Schjelderup', 'Noruega'),
    ('Sverre Nypan', 'Noruega'),
    ('Fredrik Aursnes', 'Noruega'),
    ('Diogo Costa', 'Portugal'),
    ('José Sá', 'Portugal'),
    ('Rúben Dias', 'Portugal'),
    ('Gonçalo Inácio', 'Portugal'),
    ('António Silva', 'Portugal'),
    ('Nuno Mendes', 'Portugal'),
    ('João Neves', 'Portugal'),
    ('Vitinha', 'Portugal'),
    ('Bruno Fernandes', 'Portugal'),
    ('Bernardo Silva', 'Portugal'),
    ('Rafael Leão', 'Portugal'),
    ('Pedro Neto', 'Portugal'),
    ('Gonçalo Ramos', 'Portugal'),
    ('João Félix', 'Portugal'),
    ('Cristiano Ronaldo', 'Portugal'),
    ('Rúben Neves', 'Portugal'),
    ('Rodrigo Mora', 'Portugal'),
    ('Geovany Quenda', 'Portugal'),
    ('Francisco Conceição', 'Portugal'),
    ('Angus Gunn', 'Escocia'),
    ('Craig Gordon', 'Escocia'),
    ('Andrew Robertson', 'Escocia'),
    ('Kieran Tierney', 'Escocia'),
    ('Grant Hanley', 'Escocia'),
    ('Scott McTominay', 'Escocia'),
    ('John McGinn', 'Escocia'),
    ('Billy Gilmour', 'Escocia'),
    ('Kenny McLean', 'Escocia'),
    ('Ryan Christie', 'Escocia'),
    ('Che Adams', 'Escocia'),
    ('Lawrence Shankland', 'Escocia'),
    ('Ben Doak', 'Escocia'),
    ('Tommy Conway', 'Escocia'),
    ('Lewis Ferguson', 'Escocia'),
    ('Unai Simón', 'Espanha'),
    ('David Raya', 'Espanha'),
    ('Robin Le Normand', 'Espanha'),
    ('Pau Cubarsí', 'Espanha'),
    ('Dani Vivian', 'Espanha'),
    ('Marc Cucurella', 'Espanha'),
    ('Rodri', 'Espanha'),
    ('Pedri', 'Espanha'),
    ('Gavi', 'Espanha'),
    ('Martín Zubimendi', 'Espanha'),
    ('Fabián Ruiz', 'Espanha'),
    ('Dani Olmo', 'Espanha'),
    ('Nico Williams', 'Espanha'),
    ('Lamine Yamal', 'Espanha'),
    ('Álvaro Morata', 'Espanha'),
    ('Mikel Oyarzabal', 'Espanha'),
    ('Samu Omorodion', 'Espanha'),
    ('Fermín López', 'Espanha'),
    ('Aleix García', 'Espanha'),
    ('Robin Olsen', 'Suecia'),
    ('Victor Lindelöf', 'Suecia'),
    ('Isak Hien', 'Suecia'),
    ('Emil Holm', 'Suecia'),
    ('Ludwig Augustinsson', 'Suecia'),
    ('Hugo Larsson', 'Suecia'),
    ('Dejan Kulusevski', 'Suecia'),
    ('Emil Forsberg', 'Suecia'),
    ('Jesper Karlsson', 'Suecia'),
    ('Viktor Gyökeres', 'Suecia'),
    ('Alexander Isak', 'Suecia'),
    ('Anthony Elanga', 'Suecia'),
    ('Lucas Bergvall', 'Suecia'),
    ('Roony Bardghji', 'Suecia'),
    ('Samuel Dahl', 'Suecia'),
    ('Carl Starfelt', 'Suecia'),
    ('Gregor Kobel', 'Suica'),
    ('Yann Sommer', 'Suica'),
    ('Manuel Akanji', 'Suica'),
    ('Fabian Schär', 'Suica'),
    ('Ricardo Rodríguez', 'Suica'),
    ('Silvan Widmer', 'Suica'),
    ('Granit Xhaka', 'Suica'),
    ('Denis Zakaria', 'Suica'),
    ('Remo Freuler', 'Suica'),
    ('Ardon Jashari', 'Suica'),
    ('Dan Ndoye', 'Suica'),
    ('Ruben Vargas', 'Suica'),
    ('Breel Embolo', 'Suica'),
    ('Zeki Amdouni', 'Suica'),
    ('Michel Aebischer', 'Suica'),
    ('Leon Avdullahu', 'Suica'),
    ('Nico Elvedi', 'Suica'),
    ('Mert Günok', 'Turquia'),
    ('Altay Bayındır', 'Turquia'),
    ('Merih Demiral', 'Turquia'),
    ('Abdülkerim Bardakcı', 'Turquia'),
    ('Ferdi Kadıoğlu', 'Turquia'),
    ('Kaan Ayhan', 'Turquia'),
    ('Hakan Çalhanoğlu', 'Turquia'),
    ('Orkun Kökçü', 'Turquia'),
    ('Salih Özcan', 'Turquia'),
    ('Arda Güler', 'Turquia'),
    ('Kenan Yıldız', 'Turquia'),
    ('Kerem Aktürkoğlu', 'Turquia'),
    ('Barış Alper Yılmaz', 'Turquia'),
    ('Semih Kılıçsoy', 'Turquia'),
    ('Can Uzun', 'Turquia'),
    ('Yusuf Akçiçek', 'Turquia'),
    ('İsmail Yüksek', 'Turquia'),
]

# ── Cache persistente em disco
_CACHE_PATH = SCRIPT_DIR / ".bolao_grupos_cache.pkl"

def _compute_fingerprint(paths):
    parts = []
    for p in sorted(str(x) for x in paths):
        try:
            s = Path(p).stat()
            parts.append(f"{p}|{s.st_mtime_ns}|{s.st_size}")
        except Exception:
            parts.append(f"{p}|missing")
    return hashlib.md5("\n".join(parts).encode()).hexdigest()

@st.cache_resource(show_spinner=False)
def _disk_cache_get(fingerprint):
    if not _CACHE_PATH.exists():
        return None
    try:
        with open(_CACHE_PATH, "rb") as _f:
            _c = pickle.load(_f)
        if _c.get("fp") == fingerprint:
            return _c["data"]
    except Exception:
        pass
    return None

# ══════════════════════════════════════════════════════════════════════
# LOGIN GATE
# ══════════════════════════════════════════════════════════════════════
if "auth" not in st.session_state:
    st.session_state.auth = False

if not st.session_state.auth:
    fp_b64 = img_to_b64(FRONT_PAGE) if FRONT_PAGE else ""

    st.markdown(f"""
    <style>
    section[data-testid="stSidebar"],#MainMenu,footer,header,
    [data-testid="stHeader"],[data-testid="stToolbar"]{{display:none!important;}}
    html,body{{overflow:hidden!important;height:100%!important;}}
    .stApp{{overflow:hidden!important;height:100vh!important;}}
    [data-testid="stMain"]{{overflow:hidden!important;}}

    .stApp,[data-testid="stAppViewContainer"]{{
        background: url("data:image/png;base64,{fp_b64}") left center / cover no-repeat !important;
        min-height:100vh;
    }}
    [data-testid="stMain"]>div{{background:transparent!important;}}

    .block-container{{
        padding:0 44px!important; max-width:100%!important;
        background:transparent!important;
        min-height:100vh;
    }}

    [data-testid="stHorizontalBlock"]{{
        min-height:100vh; gap:0!important;
        align-items:stretch;
    }}
    [data-testid="column"]{{min-height:100vh;}}

    [data-testid="column"]:nth-child(1){{
        background:transparent!important;
    }}

    [data-testid="column"]:nth-child(2){{
        background:rgba(6,14,26,.92)!important;
        backdrop-filter:blur(22px)!important;
        -webkit-backdrop-filter:blur(22px)!important;
        border-left:1px solid rgba(214,184,100,.22)!important;
        box-shadow:-16px 0 60px rgba(0,0,0,.6)!important;
        padding:0!important;
        display:flex!important;
        flex-direction:column!important;
        justify-content:flex-start!important;
        padding:0 44px!important;
    }}

    [data-testid="column"]:nth-child(2) input[type="password"]{{
        background:rgba(255,255,255,.07)!important;
        border:1px solid rgba(214,184,100,.35)!important;
        border-radius:10px!important; color:#fff!important;
        font-size:.95rem!important; padding:13px 15px!important;
        transition:border-color .2s,box-shadow .2s;
    }}
    [data-testid="column"]:nth-child(2) input[type="password"]:focus{{
        border-color:#D6B864!important;
        box-shadow:0 0 0 3px rgba(214,184,100,.2)!important;
    }}
    [data-testid="column"]:nth-child(2) input::placeholder{{
        color:rgba(255,255,255,.28)!important;
    }}

    [data-testid="column"]:nth-child(2) .stButton>button{{
        width:100%;
        background:#ffffff!important;
        color:#0D2B40!important; border:none!important;
        border-radius:10px!important; font-weight:800!important;
        font-size:1rem!important; letter-spacing:.3px!important;
        padding:14px 0!important; margin-top:8px!important;
        box-shadow:0 2px 12px rgba(0,0,0,.25)!important;
        transition:all .25s ease!important;
    }}
    [data-testid="column"]:nth-child(2) .stButton>button:hover{{
        background:linear-gradient(135deg,#D6B864 0%,#b89640 100%)!important;
        color:#0D2B40!important;
        transform:translateY(-2px)!important;
        box-shadow:0 8px 28px rgba(214,184,100,.55)!important;
    }}

    [data-testid="column"]:nth-child(2) [data-testid="stAlert"]{{
        background:rgba(196,30,58,.18)!important;
        border:1px solid rgba(196,30,58,.45)!important;
        border-radius:8px!important;
        color:#FCA5A5!important;
    }}
    </style>
    """, unsafe_allow_html=True)

    _, rcol = st.columns([58, 42])
    with rcol:
        st.markdown("<div style='height:42vh'></div>", unsafe_allow_html=True)
        st.markdown("""
        <div style="margin-top:0;margin-bottom:22px">
          <span style="display:inline-flex;align-items:center;gap:6px;
            background:rgba(214,184,100,.14);border:1px solid rgba(214,184,100,.4);
            border-radius:24px;padding:5px 14px;font-size:.7rem;font-weight:700;
            color:#D6B864;letter-spacing:.8px;text-transform:uppercase">
            ⚽ &nbsp;ACESSO RESTRITO
          </span>
        </div>
        <div style="font-size:2.1rem;font-weight:900;color:#fff;
                    line-height:1.15;margin-bottom:8px;letter-spacing:-.5px">
          Bolão<br><span style="color:#D6B864">Copa 2026</span>
        </div>
        <div style="font-size:.82rem;color:rgba(255,255,255,.45);margin-bottom:28px">
          EUA &nbsp;·&nbsp; México &nbsp;·&nbsp; Canadá &nbsp;·&nbsp; Jun–Jul 2026
        </div>
        <div style="height:1px;background:linear-gradient(90deg,rgba(214,184,100,.45),transparent);
                    margin-bottom:26px"></div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div style="display:inline-flex;align-items:center;gap:6px;
                    background:rgba(214,184,100,.15);border:1px solid rgba(214,184,100,.4);
                    border-radius:24px;padding:5px 14px;
                    color:#D6B864;font-size:.7rem;font-weight:700;
                    text-transform:uppercase;letter-spacing:.8px;margin-bottom:8px">
          🔑 &nbsp;SENHA DO BOLÃO
        </div>
        """, unsafe_allow_html=True)
        st.markdown("""<style>
        [data-testid="stForm"] { border: none !important; background: transparent !important; padding: 0 !important; }
        </style>""", unsafe_allow_html=True)
        with st.form("login_form", border=False):
            pwd = st.text_input("Senha", type="password",
                                placeholder="Digite a senha...",
                                label_visibility="collapsed")
            submitted = st.form_submit_button("Entrar →", width='stretch')
            if submitted:
                if pwd == _PWD:
                    st.session_state.auth = True
                    st.rerun()
                else:
                    st.error("Senha incorreta. Tente novamente.")

        st.markdown("""
        <div style="text-align:center;margin-top:30px;margin-bottom:12px;
                    font-size:.68rem;color:rgba(255,255,255,.22);letter-spacing:.6px">
          🐂 &nbsp;RUMO AO HEXA!
        </div>
        """, unsafe_allow_html=True)
    st.stop()

# ══════════════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════════════
FLAGS={
    'Mexico':'🇲🇽','Africa do Sul':'🇿🇦','Coreia do Sul':'🇰🇷','Tchequia':'🇨🇿',
    'Canada':'🇨🇦','Bosnia e Herzegovina':'🇧🇦','Qatar':'🇶🇦','Suica':'🇨🇭',
    'Brasil':'🇧🇷','Marrocos':'🇲🇦','Haiti':'🇭🇹','Escocia':'🏴󠁧󠁢󠁳󠁣󠁴󠁿',
    'Estados Unidos':'🇺🇸','Paraguai':'🇵🇾','Australia':'🇦🇺','Turquia':'🇹🇷',
    'Alemanha':'🇩🇪','Curacao':'🇨🇼','Costa do Marfim':'🇨🇮','Equador':'🇪🇨',
    'Holanda':'🇳🇱','Japao':'🇯🇵','Suecia':'🇸🇪','Tunisia':'🇹🇳',
    'Belgica':'🇧🇪','Egito':'🇪🇬','Ira':'🇮🇷','Nova Zelandia':'🇳🇿',
    'Espanha':'🇪🇸','Cabo Verde':'🇨🇻','Arabia Saudita':'🇸🇦','Uruguai':'🇺🇾',
    'Franca':'🇫🇷','Senegal':'🇸🇳','Iraque':'🇮🇶','Noruega':'🇳🇴',
    'Argentina':'🇦🇷','Argelia':'🇩🇿','Austria':'🇦🇹','Jordania':'🇯🇴',
    'Portugal':'🇵🇹','RD Congo':'🇨🇩','Uzbequistao':'🇺🇿','Colombia':'🇨🇴',
    'Inglaterra':'🏴󠁧󠁢󠁥󠁮󠁧󠁿','Croacia':'🇭🇷','Gana':'🇬🇭','Panama':'🇵🇦',
}
def F(t): return FLAGS.get(str(t),'🌍') if t and str(t) not in ('?','') else ''

COUNTRY_ISO = {
    'Mexico':'mx','Africa do Sul':'za','Coreia do Sul':'kr','Tchequia':'cz',
    'Canada':'ca','Bosnia e Herzegovina':'ba','Qatar':'qa','Suica':'ch',
    'Brasil':'br','Marrocos':'ma','Haiti':'ht','Escocia':'gb-sct',
    'Estados Unidos':'us','Paraguai':'py','Australia':'au','Turquia':'tr',
    'Alemanha':'de','Curacao':'cw','Costa do Marfim':'ci','Equador':'ec',
    'Holanda':'nl','Japao':'jp','Suecia':'se','Tunisia':'tn',
    'Belgica':'be','Egito':'eg','Ira':'ir','Nova Zelandia':'nz',
    'Espanha':'es','Cabo Verde':'cv','Arabia Saudita':'sa','Uruguai':'uy',
    'Franca':'fr','Senegal':'sn','Iraque':'iq','Noruega':'no',
    'Argentina':'ar','Argelia':'dz','Austria':'at','Jordania':'jo',
    'Portugal':'pt','RD Congo':'cd','Uzbequistao':'uz','Colombia':'co',
    'Inglaterra':'gb-eng','Croacia':'hr','Gana':'gh','Panama':'pa',
}
def FI(t):
    if not t or str(t) in ('?',''):
        return ''
    code = COUNTRY_ISO.get(str(t),'')
    if not code:
        return ''
    return f'<img src="https://flagcdn.com/16x12/{code}.png" style="vertical-align:middle;margin-right:5px;border-radius:1px" loading="lazy">'

GRP_COLORS={
    'A':'#123A56','B':'#0D2B40','C':'#0D8587','D':'#1a6b6d',
    'E':'#B2584E','F':'#8B3D35','G':'#DC884A','H':'#a86028',
    'I':'#D6B864','J':'#b89640','K':'#0D2B40','L':'#123A56',
}
FIFA_RANKINGS={
    'Mexico':16,'Africa do Sul':63,'Coreia do Sul':23,'Tchequia':38,
    'Canada':44,'Bosnia e Herzegovina':55,'Qatar':38,'Suica':20,
    'Brasil':5,'Marrocos':14,'Haiti':86,'Escocia':27,
    'Estados Unidos':11,'Paraguai':51,'Australia':23,'Turquia':36,
    'Alemanha':12,'Curacao':83,'Costa do Marfim':41,'Equador':39,
    'Holanda':8,'Japao':15,'Suecia':26,'Tunisia':30,
    'Belgica':3,'Egito':36,'Ira':25,'Nova Zelandia':98,
    'Espanha':4,'Cabo Verde':88,'Arabia Saudita':57,'Uruguai':17,
    'Franca':2,'Senegal':21,'Iraque':65,'Noruega':28,
    'Argentina':1,'Argelia':32,'Austria':24,'Jordania':86,
    'Portugal':6,'RD Congo':57,'Uzbequistao':74,'Colombia':18,
    'Inglaterra':9,'Croacia':10,'Gana':52,'Panama':72,
}
GROUPS_DATA=OrderedDict([
    ('A',['Mexico','Africa do Sul','Coreia do Sul','Tchequia']),
    ('B',['Canada','Bosnia e Herzegovina','Qatar','Suica']),
    ('C',['Brasil','Marrocos','Haiti','Escocia']),
    ('D',['Estados Unidos','Paraguai','Australia','Turquia']),
    ('E',['Alemanha','Curacao','Costa do Marfim','Equador']),
    ('F',['Holanda','Japao','Suecia','Tunisia']),
    ('G',['Belgica','Egito','Ira','Nova Zelandia']),
    ('H',['Espanha','Cabo Verde','Arabia Saudita','Uruguai']),
    ('I',['Franca','Senegal','Iraque','Noruega']),
    ('J',['Argentina','Argelia','Austria','Jordania']),
    ('K',['Portugal','RD Congo','Uzbequistao','Colombia']),
    ('L',['Inglaterra','Croacia','Gana','Panama']),
])
GL=list(GROUPS_DATA.keys())
GROUP_FIXTURES=[
    (date(2026,6,11),'A','Mexico','Africa do Sul'),(date(2026,6,11),'A','Coreia do Sul','Tchequia'),
    (date(2026,6,18),'A','Tchequia','Africa do Sul'),(date(2026,6,18),'A','Mexico','Coreia do Sul'),
    (date(2026,6,24),'A','Tchequia','Mexico'),(date(2026,6,24),'A','Africa do Sul','Coreia do Sul'),
    (date(2026,6,12),'B','Canada','Bosnia e Herzegovina'),(date(2026,6,13),'B','Qatar','Suica'),
    (date(2026,6,18),'B','Suica','Bosnia e Herzegovina'),(date(2026,6,18),'B','Canada','Qatar'),
    (date(2026,6,24),'B','Suica','Canada'),(date(2026,6,24),'B','Bosnia e Herzegovina','Qatar'),
    (date(2026,6,13),'C','Brasil','Marrocos'),(date(2026,6,13),'C','Haiti','Escocia'),
    (date(2026,6,19),'C','Escocia','Marrocos'),(date(2026,6,19),'C','Brasil','Haiti'),
    (date(2026,6,24),'C','Escocia','Brasil'),(date(2026,6,24),'C','Marrocos','Haiti'),
    (date(2026,6,12),'D','Estados Unidos','Paraguai'),(date(2026,6,13),'D','Australia','Turquia'),
    (date(2026,6,19),'D','Estados Unidos','Australia'),(date(2026,6,19),'D','Turquia','Paraguai'),
    (date(2026,6,25),'D','Turquia','Estados Unidos'),(date(2026,6,25),'D','Paraguai','Australia'),
    (date(2026,6,14),'E','Alemanha','Curacao'),(date(2026,6,14),'E','Costa do Marfim','Equador'),
    (date(2026,6,20),'E','Alemanha','Costa do Marfim'),(date(2026,6,20),'E','Equador','Curacao'),
    (date(2026,6,25),'E','Equador','Alemanha'),(date(2026,6,25),'E','Curacao','Costa do Marfim'),
    (date(2026,6,14),'F','Holanda','Japao'),(date(2026,6,14),'F','Suecia','Tunisia'),
    (date(2026,6,20),'F','Holanda','Suecia'),(date(2026,6,20),'F','Tunisia','Japao'),
    (date(2026,6,25),'F','Japao','Suecia'),(date(2026,6,25),'F','Tunisia','Holanda'),
    (date(2026,6,15),'G','Belgica','Egito'),(date(2026,6,15),'G','Ira','Nova Zelandia'),
    (date(2026,6,21),'G','Belgica','Ira'),(date(2026,6,21),'G','Nova Zelandia','Egito'),
    (date(2026,6,26),'G','Egito','Ira'),(date(2026,6,26),'G','Nova Zelandia','Belgica'),
    (date(2026,6,15),'H','Espanha','Cabo Verde'),(date(2026,6,15),'H','Arabia Saudita','Uruguai'),
    (date(2026,6,21),'H','Espanha','Arabia Saudita'),(date(2026,6,21),'H','Uruguai','Cabo Verde'),
    (date(2026,6,26),'H','Cabo Verde','Arabia Saudita'),(date(2026,6,26),'H','Uruguai','Espanha'),
    (date(2026,6,16),'I','Franca','Senegal'),(date(2026,6,16),'I','Iraque','Noruega'),
    (date(2026,6,22),'I','Franca','Iraque'),(date(2026,6,22),'I','Noruega','Senegal'),
    (date(2026,6,26),'I','Noruega','Franca'),(date(2026,6,26),'I','Senegal','Iraque'),
    (date(2026,6,16),'J','Argentina','Argelia'),(date(2026,6,16),'J','Austria','Jordania'),
    (date(2026,6,22),'J','Argentina','Austria'),(date(2026,6,22),'J','Jordania','Argelia'),
    (date(2026,6,27),'J','Argelia','Austria'),(date(2026,6,27),'J','Jordania','Argentina'),
    (date(2026,6,17),'K','Portugal','RD Congo'),(date(2026,6,17),'K','Uzbequistao','Colombia'),
    (date(2026,6,23),'K','Portugal','Uzbequistao'),(date(2026,6,23),'K','Colombia','RD Congo'),
    (date(2026,6,27),'K','Colombia','Portugal'),(date(2026,6,27),'K','RD Congo','Uzbequistao'),
    (date(2026,6,17),'L','Inglaterra','Croacia'),(date(2026,6,17),'L','Gana','Panama'),
    (date(2026,6,23),'L','Inglaterra','Gana'),(date(2026,6,23),'L','Panama','Croacia'),
    (date(2026,6,27),'L','Panama','Inglaterra'),(date(2026,6,27),'L','Croacia','Gana'),
]
MEDALS={1:'🥇',2:'🥈',3:'🥉'}

# ══════════════════════════════════════════════════════════════════════
# PONTUAÇÃO
# ══════════════════════════════════════════════════════════════════════
def sg(b1,b2,r1,r2):
    try: b1,b2,r1,r2=int(b1),int(b2),int(r1),int(r2)
    except: return 0
    if b1==r1 and b2==r2: return 5
    br=(b1>b2)-(b1<b2); rr=(r1>r2)-(r1<r2)
    if br==rr: return 3 if (b1-b2)==(r1-r2) else 2
    return 0

def calc_st(data):
    st={g:{t:{'pts':0,'played':0,'w':0,'d':0,'l':0,'gf':0,'ga':0} for t in ts}
        for g,ts in GROUPS_DATA.items()}
    for m,(_,g,t1,t2) in enumerate(GROUP_FIXTURES):
        if m not in data or data[m] is None: continue
        s1,s2=data[m]
        if s1 is None or s2 is None: continue
        try: s1,s2=int(s1),int(s2)
        except: continue
        for t,gf,ga in [(t1,s1,s2),(t2,s2,s1)]:
            st[g][t]['played']+=1; st[g][t]['gf']+=gf; st[g][t]['ga']+=ga
        if s1>s2:   st[g][t1]['pts']+=3;st[g][t1]['w']+=1;st[g][t2]['l']+=1
        elif s2>s1: st[g][t2]['pts']+=3;st[g][t2]['w']+=1;st[g][t1]['l']+=1
        else:
            st[g][t1]['pts']+=1;st[g][t1]['d']+=1
            st[g][t2]['pts']+=1;st[g][t2]['d']+=1
    return st

def sort_st(st):
    return {g:sorted(d.items(),key=lambda x:(-x[1]['pts'],-(x[1]['gf']-x[1]['ga']),-x[1]['gf'],FIFA_RANKINGS.get(x[0],150)))
            for g,d in st.items()}

def score_all(gb, bb, gr, br):
    """Pontuação: grupos (5/3/2/0) + bônus artilheiro (20)."""
    gdet = {}; gt = 0
    for m,(r1,r2) in gr.items():
        bv = gb.get(m)
        if bv is None:
            gdet[m] = None; continue
        p = sg(bv[0],bv[1],r1,r2); gdet[m] = p; gt += p
    ap = 0
    if br and bb and str(bb).strip().lower() == str(br).strip().lower():
        ap = 20
    return dict(total=gt+ap, grupos=gt, bonus=ap, art_pts=ap, gdet=gdet)

# ══════════════════════════════════════════════════════════════════════
# FILE LOADING
# ══════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def load_gab(path):
    try:
        wb=load_workbook(path,data_only=True,read_only=True); sh=wb.sheetnames
        gr={}
        if 'Jogos - Grupos' in sh:
            ws=wb['Jogos - Grupos']
            for m in range(72):
                r1=ws.cell(row=3+m,column=6).value; r2=ws.cell(row=3+m,column=7).value
                if r1 is not None and r2 is not None:
                    try: gr[m]=(int(r1),int(r2))
                    except: pass
        br=None  # artilheiro real (fase de grupos)
        if 'Apostas - Bonus' in sh:
            ws=wb['Apostas - Bonus']
            a=ws.cell(row=4,column=2).value
            if a: br=a
        return gr,br
    except Exception as e:
        st.error(f"Erro gabarito: {e}"); return {},None

@st.cache_data(show_spinner=False)
def load_part(path):
    try:
        wb=load_workbook(path,data_only=True,read_only=True); sh=wb.sheetnames
        gb={}
        if 'Apostas - Grupos' in sh:
            ws=wb['Apostas - Grupos']
            for m in range(72):
                v1=ws.cell(row=5,column=2+m*2).value; v2=ws.cell(row=5,column=3+m*2).value
                if v1 is not None and v2 is not None:
                    try: gb[m]=(int(v1),int(v2))
                    except: pass
        bb=None  # aposta artilheiro
        if 'Apostas - Bonus' in sh:
            ws=wb['Apostas - Bonus']
            bb=ws.cell(row=5,column=2).value
        return gb,bb
    except Exception:
        return {},None

@st.cache_data(show_spinner=False)
def detect():
    gab_files  = sorted(GABARITO_DIR.glob("Bolao_Copa2026_*.xlsx"))
    part_files = sorted(APOSTAS_DIR.glob("Bolao_Copa2026_*.xlsx"))
    gabs=[]; parts=[]
    for f in gab_files:
        nm = re.sub(r'(?i)^Bolao_Copa2026_', '', f.name).replace('.xlsx','').replace('_',' ').strip()
        gabs.append((nm, str(f)))
    for f in part_files:
        nm = re.sub(r'(?i)^Bolao_Copa2026_', '', f.name).replace('.xlsx','').replace('_',' ').strip()
        parts.append((nm, str(f)))
    return gabs, parts

@st.cache_data(show_spinner=False)
def load_consolidada(path):
    """Lê o arquivo consolidado e retorna {nm: {gb, bb}}."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        sh = wb.sheetnames
    except Exception as e:
        st.error(f"Erro ao ler arquivo consolidado: {e}")
        return {}
    result = {}
    if "Apostas - Grupos" in sh:
        ws = wb["Apostas - Grupos"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            nm = str(row[0]); gb = {}
            for m in range(72):
                ci = 1 + m * 2
                v1 = row[ci] if ci < len(row) else None
                v2 = row[ci + 1] if ci + 1 < len(row) else None
                if v1 is not None and v2 is not None:
                    try: gb[m] = (int(v1), int(v2))
                    except: pass
            result.setdefault(nm, {"gb": {}, "bb": None})["gb"] = gb
    if "Apostas - Bonus" in sh:
        ws = wb["Apostas - Bonus"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            nm = str(row[0])
            result.setdefault(nm, {"gb": {}, "bb": None})["bb"] = (row[1] if len(row) > 1 else None)
    return result

@st.cache_resource(show_spinner=False)
def load_all_data_consolidated(gab_path, consol_path):
    gr,br  = load_gab(gab_path)
    consol = load_consolidada(consol_path)
    bettors = []
    for nm, d in consol.items():
        gb, bb = d["gb"], d["bb"]
        sc = score_all(gb, bb, gr, br)
        bettors.append((nm, gb, bb, sc))
    bettors.sort(key=lambda x: -x[3]["total"])
    return gr, br, bettors

@st.cache_data(show_spinner=False)
def load_all_data(gab_path, parts_tuple):
    gr,br = load_gab(gab_path)
    def _load_one(item):
        nm, fp = item
        gb, bb = load_part(fp)
        sc = score_all(gb, bb, gr, br)
        return (nm, gb, bb, sc)
    n_workers = min(16, len(parts_tuple) or 1)
    with ThreadPoolExecutor(max_workers=n_workers) as ex:
        bettors = list(ex.map(_load_one, parts_tuple))
    bettors.sort(key=lambda x: -x[3]['total'])
    return gr, br, bettors

# ══════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### 🏆 Bolão Copa 2026")
    st.markdown("<div style='opacity:.6;font-size:.85rem;margin-top:-10px'>Fase de Grupos</div>",
                unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### ⚙️ Configuração")
    st.caption(f"📋 Gabarito: `{GABARITO_DIR}`")
    st.caption(f"👥 Apostas: `{APOSTAS_DIR}`")

    gabs, parts = detect()

    st.markdown("---")
    if not gabs:
        st.warning("⚠️ Sem gabarito.\nNomeie:\n`Bolao_Copa2026_Master.xlsx`")
        gab_path = None
    else:
        gsel = st.selectbox("📋 Gabarito", [n for n,_ in gabs])
        gab_path = next(p for n,p in gabs if n==gsel)
        st.success(f"✅ {gsel}")

    st.markdown("---")
    if CONSOLIDADA_PATH.exists():
        st.success("📦 Modo consolidado ativo")
        st.caption(f"`{CONSOLIDADA_PATH.name}`")
    elif parts:
        st.markdown(f"**👥 {len(parts)} participante(s):**")
        for nm,_ in parts:
            st.markdown(f"  · {nm}")
    else:
        st.warning("⚠️ Sem participantes.\nGere a planilha consolidada ou coloque apostas em `apostas/`.")

    st.markdown("---")
    if st.button("🔄 Recarregar dados", width='stretch'):
        st.cache_data.clear()
        st.cache_resource.clear()
        _CACHE_PATH.unlink(missing_ok=True)
        st.rerun()

    if MASCOTES:
        st.markdown("---")
        st.image(MASCOTES, caption="🐂 Rumo ao Hexa! 🏆", width='stretch')

    st.markdown("---")
    st.caption("Bolão Copa 2026 · Fase de Grupos · v1.0")

# ══════════════════════════════════════════════════════════════════════
# HERO HEADER
# ══════════════════════════════════════════════════════════════════════
_fav_b64 = img_to_b64(FAVICON) if FAVICON else ""
_hero_icon = (
    f'<img src="data:image/png;base64,{_fav_b64}" style="width:12rem;height:12rem;object-fit:contain">'
    if _fav_b64 else '<span style="font-size:2.5rem">⚽</span>'
)
st.markdown(f"""
<div class="hero">
  <div style="display:flex;align-items:center;gap:16px">
    <div>{_hero_icon}</div>
    <div>
      <div class="hero-title">Bolão Copa do Mundo 2026</div>
      <div class="hero-sub">Dashboard · Fase de Grupos</div>
      <div class="hero-badge">🇨🇦 Canadá &nbsp;·&nbsp; 🇲🇽 México &nbsp;·&nbsp; 🇺🇸 EUA &nbsp;·&nbsp; Junho–Julho 2026</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if not gab_path or (not CONSOLIDADA_PATH.exists() and not parts):
    st.info("👈 Configure a pasta na barra lateral. Gere a planilha consolidada ou coloque apostas em `apostas/`.")
    st.stop()

# ── Carregar dados
if CONSOLIDADA_PATH.exists():
    _all_paths = [gab_path, str(CONSOLIDADA_PATH)]
else:
    _all_paths = [gab_path] + [fp for _, fp in parts]

_fp = _compute_fingerprint(_all_paths)
_disk_data = _disk_cache_get(_fp)
if _disk_data is not None:
    gr, br, bettors = _disk_data
else:
    with st.spinner("Carregando dados..."):
        if CONSOLIDADA_PATH.exists():
            gr, br, bettors = load_all_data_consolidated(gab_path, str(CONSOLIDADA_PATH))
        else:
            gr, br, bettors = load_all_data(gab_path, tuple(parts))
    try:
        with open(_CACHE_PATH, "wb") as _cf:
            pickle.dump({"fp": _fp, "data": (gr, br, bettors)}, _cf,
                        protocol=pickle.HIGHEST_PROTOCOL)
    except Exception:
        pass
maxp = max((b[3]['total'] for b in bettors), default=1) or 1

# ── Métricas
c4 = st.columns(4)
for col,(val,lbl) in zip(c4,[
    (f"{len(gr)}/72","Jogos Grupos"),
    (str(sum(a+b for a,b in gr.values())),"Gols"),
    (str(bettors[0][3]['total']) if bettors else "0","Líder (pts)"),
    (str(len(bettors)),"Participantes"),
]):
    col.markdown(
        f'<div class="mc"><div class="mc-v">{val}</div>'
        f'<div class="mc-l">{lbl}</div></div>',
        unsafe_allow_html=True)

st.markdown("")

# ══════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════
_tab_labels = ["🏆 Ranking","⚽ Grupos","👤 Por Apostador"]
if MOSTRAR_SIMULACAO:
    _tab_labels.append("🔮 Simulação")
_tabs = st.tabs(_tab_labels)
T1, T2, T3 = _tabs[:3]
T4 = _tabs[3] if MOSTRAR_SIMULACAO else None

# ── TAB 1: RANKING ───────────────────────────────────────────────────
with T1:
    n_parts    = len(bettors)
    prize_pool = n_parts * 50              # ← ajuste o valor por pessoa aqui
    prize_pct  = {1: .70, 2: .20, 3: .10}  # ← ajuste a divisão do prêmio aqui

    _all_rk_names = [b[0] for b in bettors]
    _rc1, _rc2 = st.columns([1, 5])
    with _rc1:
        if st.button("✖ Limpar", width='stretch', key="rk_clear_ms"):
            st.session_state["rk_ms_sel"] = []
            st.rerun()
    with _rc2:
        if "rk_ms_sel" not in st.session_state:
            st.session_state["rk_ms_sel"] = []
        _rk_valid = [n for n in st.session_state["rk_ms_sel"] if n in _all_rk_names]
        _rk_chosen = st.multiselect(
            "Filtrar:", options=_all_rk_names, default=_rk_valid, key="rk_ms",
            placeholder="Selecione participantes para filtrar...",
            label_visibility="collapsed",
        )
        st.session_state["rk_ms_sel"] = _rk_chosen

    if "show_all_ranking" not in st.session_state:
        st.session_state["show_all_ranking"] = False
    _show_all = st.session_state["show_all_ranking"]

    if _rk_chosen:
        _display = [b for b in bettors if b[0] in set(_rk_chosen)]
        _use_expand = False
    else:
        _display  = bettors if _show_all else bettors[:10]
        _use_expand = True

    cA,cB = st.columns([3,2], gap="large")
    with cA:
        st.markdown('<div class="sh">🏆 Classificação Geral</div>', unsafe_allow_html=True)
        for pos,(nm,_,_,sc) in enumerate(_display,1):
            real_pos = next(i for i,(b,*_) in enumerate(bettors,1) if b==nm)
            cls   = {1:'rc1',2:'rc2',3:'rc3'}.get(real_pos,'rcN')
            medal = MEDALS.get(real_pos, f"{real_pos}°")
            pct   = int(sc['total']/maxp*100)
            prize_badge = ""
            if real_pos in prize_pct:
                val = prize_pool * prize_pct[real_pos]
                prize_badge = (f'<span style="background:rgba(214,184,100,.18);'
                    f'border:1px solid rgba(214,184,100,.4);border-radius:12px;'
                    f'padding:1px 8px;font-size:.65rem;font-weight:700;'
                    f'color:#D6B864;margin-left:6px">R$ {val:,.0f}</span>')
            st.markdown(f"""<div class="rc {cls}">
              <div style="font-size:1.35rem;width:32px;text-align:center;flex-shrink:0">{medal}</div>
              <div style="flex:1;min-width:0">
                <div class="rc-name">{nm}{prize_badge}</div>
                <div class="rc-sub">Grupos {sc['grupos']} · Bônus {sc['bonus']}</div>
                <div class="bar-bg"><div class="bar-fg" style="width:{pct}%"></div></div>
              </div>
              <div style="text-align:right;flex-shrink:0">
                <div class="rc-pts">{sc['total']}</div>
                <div class="rc-pl">pts</div>
              </div>
            </div>""", unsafe_allow_html=True)

        if _use_expand:
            if not _show_all and len(bettors) > 10:
                if st.button(f"👇 Ver todos os {len(bettors)} participantes", width='stretch'):
                    st.session_state["show_all_ranking"] = True
                    st.rerun()
            elif _show_all and len(bettors) > 10:
                if st.button("👆 Mostrar apenas top 10", width='stretch'):
                    st.session_state["show_all_ranking"] = False
                    st.rerun()

    with cB:
        st.markdown('<div class="sh">📊 Pontuação por Fase</div>', unsafe_allow_html=True)
        fig = go.Figure()
        for lbl,vals,color in [
            ('Grupos', [b[3]['grupos'] for b in _display], '#123A56'),
            ('Bônus',  [b[3]['bonus']  for b in _display], '#0D8587'),
        ]:
            fig.add_trace(go.Bar(
                name=lbl, y=[b[0] for b in _display], x=vals, orientation='h',
                marker_color=color, text=vals, textposition='inside',
                insidetextanchor='middle', textfont=dict(size=10, color='#fff')))
        fig.update_layout(
            barmode='stack', height=max(300, len(_display) * 28), margin=dict(l=0,r=5,t=5,b=5),
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            legend=dict(orientation='h', y=1.06, x=0, font=dict(size=10)),
            xaxis=dict(gridcolor='rgba(0,0,0,.08)', tickfont=dict(size=10)),
            yaxis=dict(tickfont=dict(size=10), autorange='reversed'))
        st.plotly_chart(fig, width='stretch', config={'displayModeBar':False})

        if bettors:
            nm0,_,_,sc0 = bettors[0]
            st.markdown(f"""<div class="mc" style="text-align:left;padding:14px 18px">
              <div style="font-size:.65rem;color:#5C5F62;text-transform:uppercase;letter-spacing:1px">Líder atual</div>
              <div style="font-size:1.35rem;font-weight:800;margin:3px 0">🥇 {nm0}</div>
              <div class="mc-v" style="font-size:1.9rem;text-align:left;color:#D6B864">{sc0['total']} pts</div>
              <div style="font-size:.73rem;color:#5C5F62;margin-top:3px">
                Grupos {sc0['grupos']} · Bônus {sc0['bonus']}</div>
            </div>""", unsafe_allow_html=True)

    # ── EVOLUÇÃO TEMPORAL ────────────────────────────────────────────
    st.markdown('<div class="sh">📈 Evolução de Pontuação</div>', unsafe_allow_html=True)

    CHART_COLORS = [
        '#D6B864','#0D8587','#B2584E','#2563EB',
        '#DC884A','#7B3FA0','#22C55E','#F97316',
        '#EF4444','#8B5CF6','#06B6D4','#EC4899',
    ]
    all_names = [b[0] for b in bettors]

    fc1, fc2 = st.columns([1, 5], vertical_alignment="bottom")
    with fc1:
        if st.button("👥 Todos", width='stretch'):
            st.session_state["sel_bettors"] = all_names
            st.session_state["ms_bettors"] = all_names
            st.rerun()
    with fc2:
        if "sel_bettors" not in st.session_state:
            st.session_state["sel_bettors"] = []
        valid_sel = [n for n in st.session_state["sel_bettors"] if n in all_names]
        chosen = st.multiselect(
            "Participantes:", options=all_names, default=valid_sel, key="ms_bettors",
            placeholder="Selecione participantes para comparar...",
        )
        st.session_state["sel_bettors"] = chosen

    active_bettors = [b for b in bettors if b[0] in st.session_state["sel_bettors"]]

    chart_mode = st.radio(
        "Modo:", ["📈 Evolução Acumulada", "📊 Pontuação por Fase"],
        horizontal=True, label_visibility="collapsed",
    )

    if chart_mode == "📈 Evolução Acumulada":
        all_match_dates = sorted(set(gf[0] for gf in GROUP_FIXTURES))
        fig_evo = go.Figure()
        for idx, (nm, gb, bb, sc) in enumerate(active_bettors):
            # Artilheiro lançado no último dia da fase de grupos (27/06)
            pts_by_date: dict = {date(2026, 6, 27): sc["art_pts"]}
            for m, (d_, g, t1, t2) in enumerate(GROUP_FIXTURES):
                p = sc["gdet"].get(m)
                if p is not None and p > 0:
                    pts_by_date[d_] = pts_by_date.get(d_, 0) + p
            dates_x, cumul_y = [], []
            running = 0
            for d in all_match_dates:
                running += pts_by_date.get(d, 0)
                dates_x.append(str(d)); cumul_y.append(running)
            color = CHART_COLORS[idx % len(CHART_COLORS)]
            fig_evo.add_trace(go.Scatter(
                x=dates_x, y=cumul_y, mode="lines+markers", name=nm,
                line=dict(color=color, width=2.5),
                marker=dict(size=6, color=color, line=dict(color="white", width=1.5)),
                hovertemplate=f"<b>{nm}</b><br>%{{x}}: <b>%{{y}} pts</b><extra></extra>",
            ))
        fig_evo.update_layout(
            height=400, margin=dict(l=0, r=10, t=20, b=10),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            legend=dict(orientation="h", y=1.08, x=0, font=dict(size=11), bgcolor="rgba(0,0,0,0)"),
            xaxis=dict(type="category", gridcolor="rgba(128,128,128,.10)", tickfont=dict(size=10),
                       tickangle=-35, showline=True, linecolor="rgba(128,128,128,.2)",
                       title=dict(text="Data do jogo", font=dict(size=11))),
            yaxis=dict(gridcolor="rgba(128,128,128,.12)", tickfont=dict(size=11),
                       title=dict(text="Pts acumulados", font=dict(size=11)), rangemode="tozero"),
            hovermode="x unified",
        )
        st.plotly_chart(fig_evo, width='stretch', config={"displayModeBar": False})
    else:
        PHASES_ORDER = ["Grupos","Bônus"]
        def phase_pts(sc, phase):
            if phase == "Grupos": return sc["grupos"]
            if phase == "Bônus":  return sc["bonus"]
            return 0
        sel_phase = st.radio("Fase:", PHASES_ORDER, horizontal=True, label_visibility="collapsed")
        phase_data = sorted(
            [(nm, phase_pts(sc, sel_phase)) for nm,_,_,sc in active_bettors],
            key=lambda x: -x[1],
        )
        bar_names  = [p[0] for p in phase_data]
        bar_vals   = [p[1] for p in phase_data]
        bar_colors = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(len(phase_data))]
        fig_bar = go.Figure(go.Bar(
            x=bar_names, y=bar_vals, text=bar_vals, textposition="outside",
            marker_color=bar_colors, hovertext=[p[0] for p in phase_data],
            hovertemplate="<b>%{hovertext}</b><br>%{y} pts<extra></extra>",
        ))
        fig_bar.update_layout(
            height=340, margin=dict(l=0, r=10, t=40, b=10),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            title=dict(text=f"Pontuação — fase: <b>{sel_phase}</b>", font=dict(size=13), x=0),
            xaxis=dict(gridcolor="rgba(128,128,128,.10)", tickfont=dict(size=11)),
            yaxis=dict(gridcolor="rgba(128,128,128,.12)", tickfont=dict(size=11), rangemode="tozero"),
        )
        st.plotly_chart(fig_bar, width='stretch', config={"displayModeBar": False})

# ── TAB 2: GRUPOS ────────────────────────────────────────────────────
with T2:
    st.markdown('<div class="sh">⚽ Classificação nos Grupos</div>', unsafe_allow_html=True)
    mode = st.radio("Baseado em:", ["📋 Resultados Reais", "👤 Apostas de Participante", "📊 Visão Consolidada"],
                    horizontal=True, label_visibility="collapsed")

    DOT = {1:'#22C55E',2:'#86EFAC',3:'#FB923C',4:'#F87171'}

    if mode.startswith("📊"):
        st.markdown('<div style="font-size:.8rem;opacity:.6;margin-bottom:8px">Apostas dos participantes selecionados por jogo, comparadas com o resultado real.</div>', unsafe_allow_html=True)

        _cons_all_names = [b[0] for b in bettors]
        _ca1, _ca2 = st.columns([1, 4], vertical_alignment="bottom")
        with _ca1:
            if st.button("👥 Todos", width='stretch', key="cons_todos"):
                st.session_state["cons_sel"] = _cons_all_names
                st.session_state["cons_ms"]  = _cons_all_names
                st.rerun()
        with _ca2:
            if "cons_sel" not in st.session_state:
                st.session_state["cons_sel"] = []
            _cons_valid = [n for n in st.session_state["cons_sel"] if n in _cons_all_names]
            _cons_chosen = st.multiselect(
                "Participantes:", options=_cons_all_names, default=_cons_valid,
                key="cons_ms", placeholder="Selecione participantes...",
            )
            st.session_state["cons_sel"] = _cons_chosen
        _cons_bettors = [b for b in bettors if b[0] in st.session_state["cons_sel"]] or bettors

        sel_grp = st.selectbox("Grupo:", GL, key="cons_grp")
        _cb_names = [b[0] for b in _cons_bettors]

        _max_jogo_chars = max(
            (len(t1) + 3 + len(t2))
            for _, g_, t1, t2 in GROUP_FIXTURES
            if g_ == sel_grp
        ) if any(g_ == sel_grp for _, g_, _, _ in GROUP_FIXTURES) else 20
        _jogo_col_w = max(130, _max_jogo_chars * 7 + 50)

        _STICKY_TH_G = (f"position:sticky;left:0;z-index:2;background:#0D2B40;"
                        f"text-align:left;white-space:nowrap;"
                        f"width:{_jogo_col_w}px;min-width:{_jogo_col_w}px")
        hdr_cells = "".join(
            f'<th style="white-space:nowrap;text-align:center;min-width:54px">{n}</th>'
            for n in _cb_names
        )
        hdr_html = (f'<th style="{_STICKY_TH_G}">Jogo</th>'
                    f'<th style="white-space:nowrap;min-width:46px">Data</th>'
                    f'<th style="white-space:nowrap;min-width:54px">Real</th>'
                    f'{hdr_cells}')
        rows_html = ""
        for m,(gdate,g,t1,t2) in enumerate(GROUP_FIXTURES):
            if g != sel_grp: continue
            real = gr.get(m)
            rs   = f"<b>{real[0]}–{real[1]}</b>" if real else "<span style='opacity:.4'>⏳</span>"
            ds   = gdate.strftime("%d/%m")
            cells = ""
            for nm_,bgb_,_,bsc_ in _cons_bettors:
                bet_ = bgb_.get(m)
                pts_ = bsc_['gdet'].get(m)
                if bet_ is None:
                    cells += '<td style="opacity:.3;text-align:center">—</td>'
                    continue
                bs_ = f"{bet_[0]}–{bet_[1]}"
                color_ = ('#22C55E' if pts_==5 else '#5EEAD4' if pts_==3 else
                          '#FB923C' if pts_==2 else '#F87171' if pts_==0 else 'inherit')
                cells += f'<td style="color:{color_};text-align:center">{bs_}</td>'
            _td_jogo = (f'<td style="position:sticky;left:0;z-index:1;'
                        f'background:var(--background-color,white);'
                        f'box-shadow:2px 0 4px rgba(0,0,0,.08);'
                        f'width:{_jogo_col_w}px;min-width:{_jogo_col_w}px;'
                        f'white-space:nowrap;font-size:.75rem">'
                        f'{FI(t1)}{t1} × {FI(t2)}{t2}</td>')
            rows_html += (f'<tr>{_td_jogo}'
                          f'<td style="opacity:.55;font-size:.72rem;text-align:center">{ds}</td>'
                          f'<td style="text-align:center">{rs}</td>{cells}</tr>')
        st.markdown(
            f'<div style="overflow-x:auto;-webkit-overflow-scrolling:touch;border-radius:8px">'
            f'<table class="mm-tbl" style="width:auto;table-layout:auto">'
            f'<thead><tr>{hdr_html}</tr></thead><tbody>{rows_html}</tbody>'
            f'</table></div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div style="font-size:.72rem;opacity:.5;margin-top:6px">'
            '<span style="color:#22C55E">■</span> Placar exato (5pts) &nbsp;'
            '<span style="color:#5EEAD4">■</span> Saldo exato (3pts) &nbsp;'
            '<span style="color:#FB923C">■</span> Vencedor certo (2pts) &nbsp;'
            '<span style="color:#F87171">■</span> Errou (0pts)</div>',
            unsafe_allow_html=True,
        )
    else:
        if mode.startswith("📋"):
            data_src = sort_st(calc_st(gr)) if gr else {
                g:[(t,{'pts':0,'played':0,'w':0,'d':0,'l':0,'gf':0,'ga':0}) for t in ts]
                for g,ts in GROUPS_DATA.items()}
            if not gr: st.info("Nenhum resultado preenchido no gabarito ainda.")
        else:
            sel_b = st.selectbox("Apostador", [b[0] for b in bettors], key='grp_pick')
            bd    = next(b for b in bettors if b[0]==sel_b)
            data_src = sort_st(calc_st(bd[1]))

        for row_g in [GL[i:i+3] for i in range(0,12,3)]:
            cs = st.columns(3, gap="small")
            for c,grp in zip(cs,row_g):
                with c:
                    color = GRP_COLORS.get(grp,'#123A56')
                    rows  = ""
                    for pos,(team,d) in enumerate(data_src.get(grp,[]),1):
                        gd  = d['gf']-d['ga']; pl = d['played']
                        pct = int(d['pts']/(pl*3)*100) if pl else 0
                        gdc = '#22C55E' if gd>0 else '#EF4444' if gd<0 else 'inherit'
                        gds = ('+' if gd>0 else '')+str(gd)
                        rows += f"""<tr>
                          <td><span class="dot" style="background:{DOT.get(pos,'#888')}"></span>{pos}</td>
                          <td class="nm">{FI(team)}{team}</td>
                          <td><b>{d['pts']}</b></td><td>{pl}</td>
                          <td>{d['w']}</td><td>{d['d']}</td><td>{d['l']}</td>
                          <td>{d['gf']}</td><td>{d['ga']}</td>
                          <td style="color:{gdc}">{gds}</td><td>{pct}%</td>
                        </tr>"""
                    st.markdown(f"""<div class="gb">
                      <div class="gb-hdr" style="color:{color}">Grupo {grp}</div>
                      <table class="gt">
                        <tr><th>#</th><th style="text-align:left">Seleção</th>
                            <th>P</th><th>J</th><th>V</th><th>E</th><th>D</th>
                            <th>GP</th><th>GC</th><th>SG</th><th>%</th></tr>
                        {rows}
                      </table>
                    </div>""", unsafe_allow_html=True)

# ── TAB 3: POR APOSTADOR ─────────────────────────────────────────────
with T3:
    sel = st.selectbox("👤 Apostador", [b[0] for b in bettors])
    bnm,bgb,bbb,bsc = next(b for b in bettors if b[0]==sel)
    pos   = next(i for i,(nm,*_) in enumerate(bettors,1) if nm==bnm)
    medal = MEDALS.get(pos, f"{pos}°")

    st.markdown(f"""
    <div class="hero" style="padding:18px 26px;margin-bottom:12px">
      <span style="font-size:1.5rem">{medal}</span>
      <span class="hero-title" style="font-size:1.45rem;margin-left:10px">{bnm}</span>
      <div class="hero-sub" style="margin-top:6px">
        Total: <b>{bsc['total']} pts</b> &nbsp;·&nbsp;
        Grupos: <b>{bsc['grupos']}</b> &nbsp;·&nbsp;
        Bônus: <b>{bsc['bonus']}</b>
      </div>
    </div>""", unsafe_allow_html=True)

    # Bônus — só artilheiro
    st.markdown('<div class="sh">🎁 Bônus — Artilheiro da Fase de Grupos</div>', unsafe_allow_html=True)
    ab = bbb; ar = br
    ok  = ab and ar and str(ab).strip().lower()==str(ar).strip().lower()
    ico = "✅" if ok else ("❌" if ar else "⏳")
    bvs = f"{F(ab)} {ab}" if ab else "—"
    rvs = f"{F(ar)} {ar}" if ar else "Aguardando"
    _bc1, _bc2, _bc3 = st.columns([1, 2, 1])
    with _bc2:
        st.markdown(f"""<div class="bc">
          <div class="bc-lbl">⚽ Artilheiro (20 pts)</div>
          <div class="bc-ico">{ico}</div>
          <div class="bc-bet">Apostou: {bvs}</div>
          <div class="bc-real">Real: {rvs}</div>
          <div class="bc-pts">{bsc['art_pts']} pts</div>
        </div>""", unsafe_allow_html=True)

    # Grupos jogo a jogo
    st.markdown('<div class="sh">⚽ Fase de Grupos — Jogo a Jogo</div>', unsafe_allow_html=True)
    for row_g in [GL[i:i+3] for i in range(0,12,3)]:
        gcols = st.columns(3)
        for gcol,grp in zip(gcols,row_g):
            with gcol:
                color = GRP_COLORS.get(grp,'#123A56')
                st.markdown(
                    f'<div style="font-weight:800;color:{color};font-size:.87rem;margin-bottom:5px">Grupo {grp}</div>',
                    unsafe_allow_html=True)
                html = ""
                for m,(gdate,g,t1,t2) in enumerate(GROUP_FIXTURES):
                    if g!=grp: continue
                    pts  = bsc['gdet'].get(m); bet = bgb.get(m); real = gr.get(m)
                    bs   = f"{bet[0]}–{bet[1]}" if bet else "—"
                    rs   = f"{real[0]}–{real[1]}" if real else "⏳"
                    bdg  = ('<span class="b5">5</span>' if pts==5 else
                            '<span class="b3">3</span>' if pts==3 else
                            '<span class="b2">2</span>' if pts==2 else
                            '<span class="b0">0</span>' if pts==0 else
                            '<span class="bN">–</span>')
                    ds = gdate.strftime("%d/%m")
                    html += f"""<div class="mr">
                      <div class="mr-t">{FI(t1)}{t1}<br>{FI(t2)}{t2}</div>
                      <span class="mr-s"><span style="opacity:.5;font-size:.7rem">{ds}</span> {bs} → {rs}</span>
                      {bdg}
                    </div>"""
                st.markdown(html, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
# TAB 4: SIMULAÇÃO (só fase de grupos + artilheiro)
# ══════════════════════════════════════════════════════════════════════
if MOSTRAR_SIMULACAO:
    with T4:
        for _k, _d in [("sim_rv", 0), ("sim_gr", {}), ("sim_res", None), ("sim_art", None)]:
            if _k not in st.session_state:
                st.session_state[_k] = _d
        _rv = st.session_state["sim_rv"]

        _sh1, _sh2 = st.columns([5, 1])
        with _sh1:
            st.markdown('<div class="sh">🔮 Simulador — Fase de Grupos</div>', unsafe_allow_html=True)
        with _sh2:
            if st.button("🗑 Limpar", width='stretch', key="sim_btn_clr"):
                st.session_state["sim_rv"]  += 1
                st.session_state["sim_gr"]   = {}
                st.session_state["sim_res"]  = None
                st.session_state["sim_art"]  = None
                st.rerun()

        st.caption("Preencha placares dos jogos que ainda não aconteceram e veja como o ranking ficaria. "
                   "Jogos já no gabarito ficam travados 🔒.")

        # Estado atual dos grupos (gabarito + simulação)
        _mgr_cur = {**st.session_state["sim_gr"]}
        for _m, _v in gr.items():
            _mgr_cur[_m] = _v

        _sg_sel = st.radio("Grupo:", GL, horizontal=True, label_visibility="collapsed", key="sim_grp_sel")
        _sg_games = [(m, gd, t1, t2)
                     for m, (gd, g, t1, t2) in enumerate(GROUP_FIXTURES)
                     if g == _sg_sel]
        _live_st = sort_st(calc_st(_mgr_cur))
        _grp_st  = _live_st.get(_sg_sel, [])

        _gc1, _gc2 = st.columns([5, 4])
        with _gc1:
            st.markdown(f'<div class="sh">📊 Classificação — Grupo {_sg_sel}</div>', unsafe_allow_html=True)
            _st_h = ("<table class='gt'><thead><tr>"
                     "<th></th><th style='text-align:left;min-width:110px'>Seleção</th>"
                     "<th>P</th><th>J</th><th>V</th><th>E</th><th>D</th>"
                     "<th>GP</th><th>GC</th><th>SG</th></tr></thead><tbody>")
            _st_r = ""
            for _pos, (_team, _td) in enumerate(_grp_st, 1):
                _bg  = "rgba(13,133,135,.10)" if _pos <= 2 else ""
                _fc  = "#22C55E" if _pos <= 2 else "#FB923C" if _pos == 3 else "inherit"
                _st_r += (
                    f"<tr style='background:{_bg}'>"
                    f"<td style='font-weight:700;color:{_fc}'>{_pos}</td>"
                    f"<td style='text-align:left'>{FI(_team)}{_team}</td>"
                    f"<td style='font-weight:700'>{_td['pts']}</td>"
                    f"<td>{_td['played']}</td><td>{_td['w']}</td>"
                    f"<td>{_td['d']}</td><td>{_td['l']}</td>"
                    f"<td>{_td['gf']}</td><td>{_td['ga']}</td>"
                    f"<td>{_td['gf']-_td['ga']:+d}</td></tr>")
            st.markdown(f"{_st_h}{_st_r}</tbody></table>", unsafe_allow_html=True)

        with _gc2:
            st.markdown(f'<div class="sh">⚽ Jogos — Grupo {_sg_sel}</div>', unsafe_allow_html=True)
            for _m, _gd, _t1, _t2 in _sg_games:
                _ds = _gd.strftime("%d/%m")
                if _m in gr:
                    _r = gr[_m]
                    st.markdown(
                        f'<div class="mr" style="opacity:.7">'
                        f'<div style="font-size:.68rem;opacity:.5;min-width:30px">{_ds}</div>'
                        f'<div class="mr-t">{FI(_t1)}{_t1}</div>'
                        f'<div class="mr-s">{_r[0]}–{_r[1]}</div>'
                        f'<div class="mr-t" style="text-align:right">{FI(_t2)}{_t2}</div>'
                        f'<span style="font-size:.62rem;opacity:.4;margin-left:4px">🔒</span>'
                        f'</div>',
                        unsafe_allow_html=True)
                    continue
                _cur = st.session_state["sim_gr"].get(_m, (0, 0))
                _mc1, _mc2, _mc3, _mc4, _mc5 = st.columns([0.6, 3, 0.6, 0.6, 3])
                _mc1.markdown(
                    f'<div style="font-size:.65rem;opacity:.5;padding-top:7px;text-align:right">{_ds}</div>',
                    unsafe_allow_html=True)
                _mc2.markdown(
                    f'<div style="font-size:.82rem;font-weight:700;text-align:right;'
                    f'padding-top:5px">{FI(_t1)}{_t1}</div>',
                    unsafe_allow_html=True)
                _s1 = _mc3.number_input("p1", 0, 20, int(_cur[0]),
                                        key=f"sg_{_m}_1_v{_rv}", label_visibility="collapsed")
                _s2 = _mc4.number_input("p2", 0, 20, int(_cur[1]),
                                        key=f"sg_{_m}_2_v{_rv}", label_visibility="collapsed")
                _mc5.markdown(
                    f'<div style="font-size:.82rem;font-weight:700;padding-top:5px">'
                    f'{FI(_t2)}{_t2}</div>',
                    unsafe_allow_html=True)
                st.session_state["sim_gr"][_m] = (int(_s1), int(_s2))

        # ── Simular artilheiro ────────────────────────────────────────
        with st.expander("🎯 Simular Artilheiro (Bônus · 20 pts)", expanded=False):
            if JOGADORES_ARTILHEIRO:
                _art_names = [j for j, _ in JOGADORES_ARTILHEIRO]
                _art_opts  = ["(não simular)"] + _art_names
                _cur_art   = st.session_state.get("sim_art")
                _art_idx   = (_art_opts.index(_cur_art) if _cur_art in _art_opts else 0)
                _art_choice = st.selectbox(
                    "Artilheiro", options=_art_opts, index=_art_idx,
                    key=f"sim_art_box_v{_rv}", label_visibility="collapsed",
                )
                if _art_choice != "(não simular)":
                    st.session_state["sim_art"] = _art_choice
                    _art_country = next((c for j, c in JOGADORES_ARTILHEIRO if j == _art_choice), None)
                    st.markdown(
                        f'<div style="margin-top:6px;font-size:.9rem;font-weight:600">'
                        f'{FI(_art_country) if _art_country else ""}{_art_choice}</div>',
                        unsafe_allow_html=True)
                else:
                    st.session_state["sim_art"] = None
                    st.markdown('<div style="margin-top:6px;font-size:.8rem;opacity:.4">—</div>',
                                unsafe_allow_html=True)
            else:
                st.markdown(
                    '<div style="font-size:.82rem;opacity:.5;padding:4px 0">'
                    'Lista de artilheiros ainda não configurada.</div>',
                    unsafe_allow_html=True)

        # ── Calcular ──────────────────────────────────────────────────
        st.markdown("---")
        if st.button("▶ Calcular Ranking Simulado", width='stretch', key="sim_calc"):
            _mgr_c = dict(gr)
            for _m, _vv in st.session_state["sim_gr"].items():
                if _m not in _mgr_c:
                    _mgr_c[_m] = _vv
            _sim_br = st.session_state.get("sim_art") or br
            _sim_r = []
            for _nm, _gb, _bb, _rsc in bettors:
                _sc2 = score_all(_gb, _bb, _mgr_c, _sim_br)
                _d2  = _sc2["total"] - _rsc["total"]
                _sim_r.append((_nm, _sc2["total"], _rsc["total"], _d2))
            _sim_r.sort(key=lambda x: -x[1])
            st.session_state["sim_res"] = _sim_r

        # ── Ranking resultado ─────────────────────────────────────────
        if st.session_state.get("sim_res"):
            st.markdown('<div class="sh">🏆 Ranking Simulado</div>', unsafe_allow_html=True)
            _hc = st.columns([0.5, 4, 1.5, 1.5, 1.5])
            for _c, _h in zip(_hc, ["#", "Participante", "Simulado", "Atual", "Δ"]):
                _c.markdown(
                    f'<div style="font-size:.7rem;font-weight:700;opacity:.5;'
                    f'text-transform:uppercase;letter-spacing:.8px">{_h}</div>',
                    unsafe_allow_html=True)
            for _rk, (_nm, _sp, _rp, _dlt) in enumerate(st.session_state["sim_res"], 1):
                _dc = "#22C55E" if _dlt > 0 else "#F87171" if _dlt < 0 else "inherit"
                _ds2 = f"+{_dlt}" if _dlt > 0 else str(_dlt)
                _rc2 = st.columns([0.5, 4, 1.5, 1.5, 1.5])
                _rc2[0].markdown(
                    f'<div style="font-size:.82rem;opacity:.5">{MEDALS.get(_rk, _rk)}</div>',
                    unsafe_allow_html=True)
                _rc2[1].markdown(
                    f'<div style="font-size:.85rem;font-weight:600">{_nm}</div>',
                    unsafe_allow_html=True)
                _rc2[2].markdown(
                    f'<div style="font-size:.9rem;font-weight:900;color:#D6B864">{_sp}</div>',
                    unsafe_allow_html=True)
                _rc2[3].markdown(
                    f'<div style="font-size:.85rem;opacity:.55">{_rp}</div>',
                    unsafe_allow_html=True)
                _rc2[4].markdown(
                    f'<div style="font-size:.85rem;font-weight:700;color:{_dc}">{_ds2}</div>',
                    unsafe_allow_html=True)

# ── Footer
st.markdown("---")
st.markdown(
    "<div style='text-align:center;color:#5C5F62;font-size:.8rem;padding:.5rem 0'>"
    "Bolão Copa do Mundo 2026 &nbsp;·&nbsp; Fase de Grupos &nbsp;·&nbsp; 🐂 Rumo ao Hexa!</div>",
    unsafe_allow_html=True)